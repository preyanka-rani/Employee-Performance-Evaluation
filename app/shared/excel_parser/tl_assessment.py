"""
app/shared/excel_parser/tl_assessment.py
─────────────────────────────────────────
Legacy TL Assessment upload parser.

Supports the minimal 4-column upload format used by ``POST /api/v1/upload/tl-scores``:
    | employee_email | problem_solving | kpi | general |

This parser is intentionally separate from the team-aware
``app.shared.excel_parser.parser.parse_tl_excel`` (used by the bulk-run
endpoint) because the upload flow has a different contract:
synchronous, returns ``TLAssessmentRow`` (legacy schema), and is
decoupled from any specific team's data shape.
"""

from __future__ import annotations

import io
from dataclasses import dataclass

import openpyxl
from pydantic import ValidationError

from app.core.logging_config import get_logger
from app.schemas.scores import TLAssessmentRow
from app.shared.excel_parser.parser import ExcelParseError

logger = get_logger(__name__)


# Mapping from human-friendly column names to normalised keys
_COLUMN_MAP: dict[str, str] = {
    "employee_email": "employee_email",
    "email": "employee_email",
    "problem_solving": "problem_solving",
    "problem solving": "problem_solving",
    "critical thinking": "problem_solving",
    "kpi": "kpi",
    "performance agreement": "kpi",
    "general": "general",
    "general assessment": "general",
    "team lead general assessment": "general",
}

_REQUIRED_KEYS: set[str] = {"employee_email", "problem_solving", "kpi", "general"}


@dataclass
class ExcelParseResult:
    rows: list[TLAssessmentRow]
    errors: list[str]


def parse_tl_assessment_excel(file_bytes: bytes) -> ExcelParseResult:
    """
    Parse a TL Assessment Excel upload.

    Args:
        file_bytes: Raw bytes of the uploaded .xlsx / .xls file.

    Returns:
        ExcelParseResult with valid rows and any per-row error messages.
    """
    errors: list[str] = []
    rows: list[TLAssessmentRow] = []

    try:
        workbook = openpyxl.load_workbook(
            io.BytesIO(file_bytes), read_only=True, data_only=True
        )
    except Exception as exc:
        raise ExcelParseError([f"Cannot open Excel file: {exc}"]) from exc

    sheet = workbook.active
    if sheet is None:
        raise ExcelParseError(["Excel workbook has no active sheet."])

    # ── Read header row ───────────────────────────────────────────────────────
    header_row = next(sheet.iter_rows(min_row=1, max_row=1, values_only=True), None)
    if header_row is None:
        raise ExcelParseError(["Excel file appears to be empty."])

    # Normalise header names
    col_index: dict[str, int] = {}
    for idx, cell_value in enumerate(header_row):
        if cell_value is None:
            continue
        normalised = _COLUMN_MAP.get(str(cell_value).strip().lower())
        if normalised:
            col_index[normalised] = idx

    missing_cols = _REQUIRED_KEYS - col_index.keys()
    if missing_cols:
        raise ExcelParseError(
            [f"Missing required columns: {', '.join(sorted(missing_cols))}"]
        )

    # ── Parse data rows ───────────────────────────────────────────────────────
    for row_num, row in enumerate(
        sheet.iter_rows(min_row=2, values_only=True), start=2
    ):
        if all(cell is None for cell in row):
            continue  # skip entirely blank rows

        def _get(key: str) -> str:
            val = row[col_index[key]]
            return str(val).strip() if val is not None else ""

        raw = {
            "employee_email": _get("employee_email").lower(),
            "problem_solving": _get("problem_solving"),
            "kpi": _get("kpi"),
            "general": _get("general"),
        }

        try:
            validated = TLAssessmentRow(**raw)
            rows.append(validated)
        except ValidationError as exc:
            for err in exc.errors():
                field_name = ".".join(str(loc) for loc in err["loc"])
                errors.append(f"Row {row_num}, field '{field_name}': {err['msg']}")
        except Exception as exc:
            errors.append(f"Row {row_num}: unexpected error – {exc}")

    workbook.close()
    logger.info("excel_parsed", valid_rows=len(rows), error_count=len(errors))
    return ExcelParseResult(rows=rows, errors=errors)
