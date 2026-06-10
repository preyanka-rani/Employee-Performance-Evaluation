"""
app/services/reporting/report_generator.py
────────────────────────────────────────────
Compiles all individual score rows into structured report responses.

This module does not calculate scores — it only reads from the DB and
assembles the response schemas for the report endpoints.
"""

from __future__ import annotations

from sqlalchemy.ext.asyncio import AsyncSession

from app.repositories.employee_repository import EmployeeRepository
from app.repositories.score_repository import (
    AttendanceRepository,
    CodeQualityRepository,
    DeveloperFinalScoreRepository,
    FinalScoreRepository,
    SentimentRepository,
    TLAssessmentRepository,
    WorkLogRepository,
)
from app.schemas.reports import (
    EmployeeReportResponse,
    TeamReportEntry,
    TeamReportResponse,
)


async def generate_employee_report(
    employee_id: str,
    year: int,
    month: int,
    db: AsyncSession,
) -> EmployeeReportResponse | None:
    """
    Assemble the full score breakdown for one employee.

    Returns None if the employee or their evaluation data doesn't exist.
    """
    emp_repo = EmployeeRepository(db)
    employee = await emp_repo.get_by_employee_id(employee_id)
    if employee is None:
        return None

    fs_repo = FinalScoreRepository(db)
    score = await fs_repo.get_by_email_and_period(
        email=employee.email,
        year=year,
        month=month,
    )
    if score is None:
        return None

    return EmployeeReportResponse(
        employee_id=employee.employee_id,
        name=employee.name,
        email=employee.email,
        team=employee.team,
        year=year,
        month=month,
        quality_check_score=score.quality_check_score,
        work_log_score=score.work_log_score,
        sentiment_score=score.sentiment_score,
        attendance_score=score.attendance_score,
        problem_solving_score=score.problem_solving_score,
        kpi_score=score.kpi_score,
        general_score=score.general_score,
        segment_a_marks=score.segment_a_marks,
        segment_b_marks=score.segment_b_marks,
        base_total=score.base_total,
        reward_score=score.reward_score,
        final_score=score.final_score,
    )


async def generate_team_report(
    team: str,
    year: int,
    month: int,
    db: AsyncSession,
) -> TeamReportResponse | None:
    """
    Assemble the team report for all active employees in a team.

    Returns None if no employees exist for the team.
    """
    emp_repo = EmployeeRepository(db)
    employees = await emp_repo.get_by_team(team)
    if not employees:
        return None

    emails = [e.email for e in employees]
    email_to_emp = {e.email: e for e in employees}

    fs_repo = FinalScoreRepository(db)
    scores = await fs_repo.get_team_scores_by_period(
        emails=emails,
        year=year,
        month=month,
    )

    entries: list[TeamReportEntry] = []
    for score in scores:
        emp = email_to_emp.get(score.employee_email)
        entries.append(
            TeamReportEntry(
                employee_id=emp.employee_id if emp else "unknown",
                name=emp.name if emp else "unknown",
                email=score.employee_email,
                final_score=score.final_score,
                segment_a_marks=score.segment_a_marks,
                segment_b_marks=score.segment_b_marks,
                base_total=score.base_total,
                reward_score=score.reward_score,
            )
        )

    entries.sort(key=lambda e: e.final_score, reverse=True)
    team_avg = (
        round(sum(e.final_score for e in entries) / len(entries), 2) if entries else 0.0
    )

    return TeamReportResponse(
        team=team,
        year=year,
        month=month,
        employee_count=len(entries),
        team_average_score=team_avg,
        entries=entries,
    )


async def generate_excel_report(
    run_id: int,
    emails: list[str],
    team: str,
    year: int,
    month: int,
    db: AsyncSession,
) -> str:
    """
    Build a formatted Excel report for a completed bulk evaluation run.

    For the developer team this reads from ``developer_final_scores`` so
    every sub-component column is available without extra joins.

    Columns (24 total):
        Employee ID | Name | Email |
        Component 1 Score | Code Quality (30%) | Resolution Rate % (35%) |
        Reopen Quality (15%) | Lines Added Score (10%) | Lines Deleted Score (10%) |
        Work Log Hours | Work Log Score | Sentiment Score | Component 2 Score |
        Attendance Score |
        TL Problem Solving | TL KPI | TL General | TL Total |
        Segment A Marks | Segment B Marks |
        Base Total | Reward Score | Final Score

    Color coding on Final Score column:
        ≥ 80  → green fill
        60–79 → yellow fill
        < 60  → red fill

    Saves to:  outputs/developer/Final_Report_{team}_{year}_{month:02d}.xlsx
    Returns the saved file path as a string.
    """
    import pathlib

    import openpyxl
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils import get_column_letter

    dev_repo = DeveloperFinalScoreRepository(db)
    scores = await dev_repo.get_by_run_id(run_id=run_id, emails=emails)

    # ── Workbook setup ────────────────────────────────────────────────────────
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = f"{team} {year}-{month:02d}"

    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill("solid", fgColor="1F4E79")
    header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    green_fill = PatternFill("solid", fgColor="C6EFCE")
    yellow_fill = PatternFill("solid", fgColor="FFEB9C")
    red_fill = PatternFill("solid", fgColor="FFC7CE")

    headers = [
        "Employee ID",
        "Name",
        "Email",
        # Component 1
        "Component 1 Score (0-100)",
        "  Code Quality (30%)",
        "  Resolution Rate % (35%)",
        "  Reopen Quality (15%)",
        "  Lines Added Score (10%)",
        "  Lines Deleted Score (10%)",
        # Component 2
        "Work Log Hours",
        "Work Log Score (0-100)",
        "Sentiment Score (0-100)",
        "Component 2 Score (0-100)",
        # Segment B inputs
        "Attendance Score (0-100)",
        "TL Problem Solving (0-10)",
        "TL KPI (0-15)",
        "TL General (0-15)",
        "TL Total (0-40)",
        # Totals
        "Segment A Marks (0-50)",
        "Segment B Marks (0-50)",
        "Base Total (0-100)",
        "Reward Score (0-5)",
        "Final Score",
    ]
    ws.append(headers)

    # Style header row
    for col_idx, _ in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col_idx)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment

    ws.freeze_panes = "A2"

    # ── Data rows ─────────────────────────────────────────────────────────────
    for s in sorted(scores, key=lambda x: x.final_score, reverse=True):
        row_data = [
            s.employee_id,
            s.employee_name,
            s.employee_email,
            round(s.component1_score, 2),
            round(s.code_quality_score, 2),
            round(s.resolution_rate, 2),
            round(s.reopen_quality_score, 2),
            round(s.lines_added_score, 2),
            round(s.lines_deleted_score, 2),
            round(s.work_log_hours, 2),
            round(s.work_log_score, 2),
            round(s.sentiment_score, 2),
            round(s.component2_score, 2),
            round(s.attendance_score, 2),
            round(s.tl_problem_solving, 2),
            round(s.tl_kpi, 2),
            round(s.tl_general, 2),
            round(s.tl_total, 2),
            round(s.segment_a_marks, 2),
            round(s.segment_b_marks, 2),
            round(s.base_total, 2),
            round(s.reward_score, 2),
            round(s.final_score, 2),
        ]
        ws.append(row_data)

        # Color-code Final Score cell (last column)
        final_cell = ws.cell(row=ws.max_row, column=len(headers))
        if s.final_score >= 80:
            final_cell.fill = green_fill
        elif s.final_score >= 60:
            final_cell.fill = yellow_fill
        else:
            final_cell.fill = red_fill

    # ── Column widths ─────────────────────────────────────────────────────────
    column_widths = [
        14,
        22,
        30,  # id / name / email
        24,
        22,
        24,
        22,
        22,
        22,  # comp1 and sub-scores
        18,
        22,
        22,
        24,  # comp2 block
        22,  # attendance
        24,
        14,
        18,
        14,  # TL scores
        22,
        22,  # segment marks
        20,
        18,
        15,  # final cols
    ]
    for col_idx, width in enumerate(column_widths, start=1):
        ws.column_dimensions[get_column_letter(col_idx)].width = width

    # ── Summary row ───────────────────────────────────────────────────────────
    if scores:
        avg_final = round(sum(s.final_score for s in scores) / len(scores), 2)
        ws.append([""] * (len(headers) - 2) + ["Team Average", avg_final])
        summary_row = ws.max_row
        for col_idx in range(1, len(headers) + 1):
            ws.cell(row=summary_row, column=col_idx).font = Font(bold=True)

    # ── Save file ─────────────────────────────────────────────────────────────
    output_dir = pathlib.Path("outputs") / team
    output_dir.mkdir(parents=True, exist_ok=True)
    filename = f"Final_Report_{team}_{year}_{month:02d}.xlsx"
    output_path = output_dir / filename
    wb.save(str(output_path))

    return str(output_path)


def _parse_bundle_ref(mr_reference: str) -> tuple[str, int]:
    """
    Parse ``analysis_reference`` back into (project_path, commit_count).

    Format: "namespace/project (2026-01, 34 commits)"
    Returns ("namespace/project", 34).  Falls back gracefully on unexpected input.
    """
    try:
        if " (" in mr_reference and "commits)" in mr_reference:
            project_path = mr_reference.split(" (")[0].strip()
            inside = mr_reference.split(" (")[1].rstrip(")")
            # inside looks like "2026-01, 34 commits"
            commit_part = inside.split(", ")[-1]  # "34 commits"
            commit_count = int(commit_part.split()[0])
            return project_path, commit_count
    except (IndexError, ValueError):
        pass
    return mr_reference, 0


async def generate_code_quality_report(
    run_id: int,
    emails: list[str],
    team: str,
    year: int,
    month: int,
    db: AsyncSession,
) -> str:
    """
    Build a detailed per-project code quality report for a completed bulk run.

    Sheet 1 — "Summary"
        One row per developer.  Weighted-average score, total lines added/deleted,
        and a count of doc-only skipped bundles.

    Sheet 2 — "Per-Project Analysis"
        One row per project per developer (scored bundles first, then skipped).
        Columns: Developer | Employee ID | Project | Commits | Lines Added |
                 Lines Deleted | Net Lines | Readability | Logic | Error Handling |
                 Architecture | Security | AI Score | Score Weight |
                 Developer Final Score | Issues Identified | AI Remarks

    Color coding (score columns):
        ≥ 80 → green  |  60–79 → yellow  |  < 60 → red  |  skipped → grey

    Saves to:  outputs/developer/CodeQuality_Report_{team}_{year}_{month:02d}.xlsx
    Returns the saved file path.
    """
    import json
    import pathlib

    import openpyxl
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
    from openpyxl.utils import get_column_letter

    # ── Fetch data ────────────────────────────────────────────────────────────
    cq_repo = CodeQualityRepository(db)
    cq_rows = await cq_repo.get_by_run_id_and_emails(run_id=run_id, emails=emails)

    emp_repo = EmployeeRepository(db)
    employees = {e.email: e for e in await emp_repo.get_by_team(team)}

    from collections import defaultdict

    rows_by_email: dict[str, list] = defaultdict(list)
    for row in cq_rows:
        rows_by_email[row.employee_email].append(row)

    # ── Style helpers ─────────────────────────────────────────────────────────
    wb = openpyxl.Workbook()

    hdr_font = Font(bold=True, color="FFFFFF", size=10)
    dev_font = Font(bold=True, color="1F4E79", size=10)
    data_font = Font(size=9)
    nocommit_font = Font(italic=True, color="7F7F7F", size=9)
    skip_font = Font(italic=True, color="999999", size=9)

    hdr_fill = PatternFill("solid", fgColor="1F4E79")
    dev_fill_a = PatternFill("solid", fgColor="D6E4F0")
    dev_fill_b = PatternFill("solid", fgColor="EBF5FB")
    nocommit_fill = PatternFill("solid", fgColor="F2F2F2")
    skip_fill = PatternFill("solid", fgColor="EEEEEE")
    orange_fill = PatternFill("solid", fgColor="FCE4D6")  # AI-failed rows
    failed_font = Font(italic=True, color="C55A11", size=9)
    green_fill = PatternFill("solid", fgColor="C6EFCE")
    yellow_fill = PatternFill("solid", fgColor="FFEB9C")
    red_fill = PatternFill("solid", fgColor="FFC7CE")
    blue_fill = PatternFill("solid", fgColor="BDD7EE")

    center = Alignment(horizontal="center", vertical="center", wrap_text=True)
    left_va = Alignment(horizontal="left", vertical="center", wrap_text=False)
    wrap = Alignment(horizontal="left", vertical="top", wrap_text=True)

    thin_side = Side(border_style="thin", color="BDD7EE")
    thin_border = Border(
        left=thin_side, right=thin_side, top=thin_side, bottom=thin_side
    )
    thick_bottom = Border(
        left=thin_side,
        right=thin_side,
        top=thin_side,
        bottom=Side(border_style="medium", color="1F4E79"),
    )

    def _score_fill(score: float) -> PatternFill:
        if score >= 80:
            return green_fill
        if score >= 60:
            return yellow_fill
        return red_fill

    def _apply_borders(ws, row: int, ncols: int, border: Border = thin_border) -> None:
        for c in range(1, ncols + 1):
            ws.cell(row=row, column=c).border = border

    def _set_row_fill_font(
        ws, row: int, ncols: int, fill: PatternFill, font: Font
    ) -> None:
        for c in range(1, ncols + 1):
            ws.cell(row=row, column=c).fill = fill
            ws.cell(row=row, column=c).font = font

    # ══════════════════════════════════════════════════════════════════════════
    # Sheet 1 — Summary  (11 columns)
    # ══════════════════════════════════════════════════════════════════════════
    SUM_COLS = 11
    ws_sum = wb.active
    ws_sum.title = "Summary"

    ws_sum.merge_cells(f"A1:{get_column_letter(SUM_COLS)}1")
    tc = ws_sum["A1"]
    tc.value = f"CODE QUALITY SUMMARY  —  {team.upper()}  ({year}-{month:02d})"
    tc.font = Font(bold=True, color="FFFFFF", size=13)
    tc.fill = hdr_fill
    tc.alignment = Alignment(horizontal="center", vertical="center")
    ws_sum.row_dimensions[1].height = 28

    sum_headers = [
        "Employee ID",
        "Developer Name",
        "Email",
        "GitLab Username",
        "Projects\nAnalyzed",
        "Doc-only\nSkipped",
        "Total\nCommits",
        "Total Lines\nAdded",
        "Total Lines\nDeleted",
        "Weighted Avg\nScore (0–100)",
        "Status",
    ]
    for ci, h in enumerate(sum_headers, 1):
        cell = ws_sum.cell(row=2, column=ci, value=h)
        cell.font = hdr_font
        cell.fill = hdr_fill
        cell.alignment = center
    ws_sum.freeze_panes = "A3"
    ws_sum.row_dimensions[2].height = 36

    sum_col_widths = [13, 24, 34, 20, 14, 14, 14, 15, 15, 20, 38]
    for i, w in enumerate(sum_col_widths, 1):
        ws_sum.column_dimensions[get_column_letter(i)].width = w

    for ei, email in enumerate(emails):
        emp = employees.get(email)
        emp_id = emp.employee_id if emp else "N/A"
        emp_name = emp.name if emp else email
        gitlab_u = (emp.gitlab_username or emp.employee_id) if emp else "N/A"

        bundles = rows_by_email.get(email, [])
        no_com_b = [b for b in bundles if b.mr_reference == "no_commits_found"]
        skipped_b = [
            b
            for b in bundles
            if b.mr_reference != "no_commits_found"
            and getattr(b, "model_used", "") == "doc_config_skipped"
        ]
        failed_b = [
            b
            for b in bundles
            if b.mr_reference != "no_commits_found"
            and getattr(b, "model_used", "") == "ai_failed"
        ]
        real_b = [
            b
            for b in bundles
            if b.mr_reference != "no_commits_found"
            and getattr(b, "model_used", "") not in ("doc_config_skipped", "ai_failed")
        ]

        # Include ALL project bundles (scored + skipped + failed) in line/commit totals
        total_lines_added = sum(
            getattr(b, "lines_added", 0) for b in real_b + skipped_b + failed_b
        )
        total_lines_deleted = sum(
            getattr(b, "lines_deleted", 0) for b in real_b + skipped_b + failed_b
        )

        total_commits_sum = 0
        avg_score: float | None = None
        score_display: float | str = "—"

        if real_b:
            weighted_sum = 0.0
            for b in real_b:
                _, cnt = _parse_bundle_ref(b.mr_reference)
                total_commits_sum += cnt
                weighted_sum += cnt * b.raw_score
            # Also count commits from failed bundles for the total display
            for b in failed_b:
                _, cnt = _parse_bundle_ref(b.mr_reference)
                total_commits_sum += cnt
            avg_score = (
                round(
                    weighted_sum
                    / sum(_parse_bundle_ref(b.mr_reference)[1] for b in real_b),
                    2,
                )
                if real_b
                else None
            )
            score_display = avg_score
            if failed_b:
                status = f"✔ Analyzed ({len(real_b)}) ⚠ AI Failed ({len(failed_b)})"
            else:
                status = "✔ Analyzed"
        elif failed_b:
            for b in failed_b:
                _, cnt = _parse_bundle_ref(b.mr_reference)
                total_commits_sum += cnt
            status = f"⚠ AI Failed — {len(failed_b)} project(s), commits tracked"
        elif no_com_b:
            status = "✘ No Commits Found"
        else:
            status = "No Data"

        ws_sum.append(
            [
                emp_id,
                emp_name,
                email,
                gitlab_u,
                len(real_b),
                len(skipped_b),
                total_commits_sum,
                total_lines_added,
                total_lines_deleted,
                score_display,
                status,
            ]
        )
        dr = ws_sum.max_row
        row_fill = dev_fill_a if ei % 2 == 0 else dev_fill_b
        _set_row_fill_font(ws_sum, dr, SUM_COLS, row_fill, data_font)
        _apply_borders(ws_sum, dr, SUM_COLS)

        # Score cell (col 10) — color coded
        sc = ws_sum.cell(row=dr, column=10)
        sc.alignment = center
        sc.font = Font(bold=True, size=10)
        if avg_score is not None:
            sc.fill = _score_fill(avg_score)
        elif failed_b and not no_com_b:
            sc.fill = orange_fill
            sc.font = Font(italic=True, color="C00000", size=10)
        else:
            sc.fill = nocommit_fill
            sc.font = Font(italic=True, color="7F7F7F", size=10)

        # Lines added/deleted cols — light blue tint (blue even if 0, shows data exists)
        for col_lines in (8, 9):
            ws_sum.cell(row=dr, column=col_lines).alignment = center
            if total_lines_added > 0 or total_lines_deleted > 0:
                ws_sum.cell(row=dr, column=col_lines).fill = blue_fill

        ws_sum.row_dimensions[dr].height = 22

    # ══════════════════════════════════════════════════════════════════════════
    # Sheet 2 — Per-Project Analysis  (17 columns)
    # ══════════════════════════════════════════════════════════════════════════
    DET_COLS = 17
    ws_det = wb.create_sheet("Per-Project Analysis")

    ws_det.merge_cells(f"A1:{get_column_letter(DET_COLS)}1")
    tc2 = ws_det["A1"]
    tc2.value = (
        f"PER-PROJECT CODE QUALITY & COMMIT REPORT"
        f"  —  {team.upper()}  ({year}-{month:02d})"
    )
    tc2.font = Font(bold=True, color="FFFFFF", size=13)
    tc2.fill = hdr_fill
    tc2.alignment = Alignment(horizontal="center", vertical="center")
    ws_det.row_dimensions[1].height = 28

    det_headers = [
        "Developer Name",  # 1
        "Employee ID",  # 2
        "Project / Repository",  # 3
        "Total\nCommits",  # 4
        "Lines\nAdded",  # 5  ← NEW
        "Lines\nDeleted",  # 6  ← NEW
        "Net\nLines",  # 7  ← NEW
        "Readability\n(0–100)",  # 8
        "Logic &\nEfficiency",  # 9
        "Error\nHandling",  # 10
        "Architecture",  # 11
        "Security",  # 12
        "Project AI\nScore (0–100)",  # 13
        "Score Weight\n(Commits×Score)",  # 14
        "Developer Final\nScore (Wt. Avg)",  # 15
        "Issues\nIdentified",  # 16
        "AI Remarks",  # 17
    ]
    for ci, h in enumerate(det_headers, 1):
        cell = ws_det.cell(row=2, column=ci, value=h)
        cell.font = hdr_font
        cell.fill = hdr_fill
        cell.alignment = center
    ws_det.freeze_panes = "A3"
    ws_det.row_dimensions[2].height = 42

    det_col_widths = [
        24,
        13,
        36,
        12,
        12,
        12,
        10,
        13,
        14,
        13,
        14,
        11,
        16,
        18,
        18,
        45,
        75,
    ]
    for i, w in enumerate(det_col_widths, 1):
        ws_det.column_dimensions[get_column_letter(i)].width = w

    for ei, email in enumerate(emails):
        emp = employees.get(email)
        emp_name = emp.name if emp else email
        emp_id = emp.employee_id if emp else "N/A"

        bundles = rows_by_email.get(email, [])
        no_com_b = [b for b in bundles if b.mr_reference == "no_commits_found"]
        skipped_b = [
            b
            for b in bundles
            if b.mr_reference != "no_commits_found"
            and getattr(b, "model_used", "") == "doc_config_skipped"
        ]
        failed_b = [
            b
            for b in bundles
            if b.mr_reference != "no_commits_found"
            and getattr(b, "model_used", "") == "ai_failed"
        ]
        real_b = [
            b
            for b in bundles
            if b.mr_reference != "no_commits_found"
            and getattr(b, "model_used", "") not in ("doc_config_skipped", "ai_failed")
        ]

        row_bg = dev_fill_a if ei % 2 == 0 else dev_fill_b

        # ── No data at all ────────────────────────────────────────────────────
        if not bundles:
            ws_det.append(
                [
                    emp_name,
                    emp_id,
                    "No data found for this employee",
                    0,
                    0,
                    0,
                    0,
                    "—",
                    "—",
                    "—",
                    "—",
                    "—",
                    "—",
                    "—",
                    "—",
                    "No records found",
                    f"Re-run evaluation for run #{run_id}",
                ]
            )
            dr = ws_det.max_row
            _set_row_fill_font(ws_det, dr, DET_COLS, nocommit_fill, nocommit_font)
            _apply_borders(ws_det, dr, DET_COLS)
            ws_det.row_dimensions[dr].height = 30
            continue

        # ── No commits ────────────────────────────────────────────────────────
        if no_com_b and not real_b and not skipped_b and not failed_b:
            sentinel = no_com_b[0]
            ws_det.append(
                [
                    emp_name,
                    emp_id,
                    "— No commits found in this period —",
                    0,
                    0,
                    0,
                    0,
                    "—",
                    "—",
                    "—",
                    "—",
                    "—",
                    "—",
                    "—",
                    "—",
                    "—",
                    sentinel.reasoning,
                ]
            )
            dr = ws_det.max_row
            _set_row_fill_font(ws_det, dr, DET_COLS, nocommit_fill, nocommit_font)
            ws_det.cell(row=dr, column=17).alignment = wrap
            _apply_borders(ws_det, dr, DET_COLS, thick_bottom)
            ws_det.row_dimensions[dr].height = 55
            continue

        # Developer-level weighted final score (from real/scored bundles only)
        total_commits_all = 0
        weighted_score_all = 0.0
        for b in real_b:
            _, cnt = _parse_bundle_ref(b.mr_reference)
            total_commits_all += cnt
            weighted_score_all += cnt * b.raw_score
        dev_final = (
            round(weighted_score_all / total_commits_all, 2)
            if total_commits_all > 0
            else (
                round(sum(b.raw_score for b in real_b) / len(real_b), 2)
                if real_b
                else 0.0
            )
        )

        all_display_bundles = (
            real_b + skipped_b + failed_b
        )  # scored, skipped, then failed
        for bi, bundle in enumerate(all_display_bundles):
            project_path, commit_count = _parse_bundle_ref(bundle.mr_reference)
            is_skipped_row = getattr(bundle, "model_used", "") == "doc_config_skipped"
            is_failed_row = getattr(bundle, "model_used", "") == "ai_failed"

            lines_added_val = getattr(bundle, "lines_added", 0)
            lines_deleted_val = getattr(bundle, "lines_deleted", 0)
            net_lines_val = lines_added_val - lines_deleted_val

            if is_skipped_row or is_failed_row:
                score_weight_val: float | str = "—"
                raw_score_display: float | str = "—"
                sub_scores: list = ["—"] * 5
            else:
                score_weight_val = round(commit_count * bundle.raw_score, 1)
                raw_score_display = round(bundle.raw_score, 2)
                sub_scores = [
                    round(bundle.readability_score, 1),
                    round(bundle.logic_efficiency_score, 1),
                    round(bundle.error_handling_score, 1),
                    round(bundle.architecture_score, 1),
                    round(bundle.security_score, 1),
                ]

            try:
                issues_list = json.loads(bundle.issues) if bundle.issues else []
                issues_str = (
                    "\n".join(f"• {i}" for i in issues_list)
                    if issues_list
                    else (
                        "Doc/Config only — skipped"
                        if is_skipped_row
                        else (
                            "⚠ AI analysis failed — API key invalid / rate limit exceeded"
                            if is_failed_row
                            else "None identified"
                        )
                    )
                )
            except (json.JSONDecodeError, TypeError):
                issues_str = bundle.issues or "None"

            dev_name_cell = emp_name if bi == 0 else ""
            dev_id_cell = emp_id if bi == 0 else ""
            final_cell: float | str = dev_final if (bi == 0 and real_b) else ""

            ws_det.append(
                [
                    dev_name_cell,
                    dev_id_cell,
                    project_path,
                    commit_count,
                    lines_added_val,
                    lines_deleted_val,
                    net_lines_val,
                    sub_scores[0],
                    sub_scores[1],
                    sub_scores[2],
                    sub_scores[3],
                    sub_scores[4],
                    raw_score_display,
                    score_weight_val,
                    final_cell,
                    issues_str,
                    bundle.reasoning,
                ]
            )
            dr = ws_det.max_row

            if is_failed_row:
                _set_row_fill_font(ws_det, dr, DET_COLS, orange_fill, failed_font)
                ws_det.cell(row=dr, column=1).font = (
                    Font(bold=True, italic=True, color="C55A11", size=9)
                    if bi == 0
                    else failed_font
                )
            elif is_skipped_row:
                _set_row_fill_font(ws_det, dr, DET_COLS, skip_fill, skip_font)
            else:
                _set_row_fill_font(ws_det, dr, DET_COLS, row_bg, data_font)
                ws_det.cell(row=dr, column=1).font = dev_font if bi == 0 else data_font

                # Color the AI score cell
                ws_det.cell(row=dr, column=13).fill = _score_fill(bundle.raw_score)
                ws_det.cell(row=dr, column=13).font = Font(bold=True, size=9)
                ws_det.cell(row=dr, column=14).alignment = center

                # Developer final score (first row only)
                if bi == 0 and real_b:
                    fc = ws_det.cell(row=dr, column=15)
                    fc.fill = _score_fill(dev_final)
                    fc.font = Font(bold=True, size=9)
                    fc.alignment = center

            # Lines added/deleted — blue tint for readability (real and failed rows)
            for col_l in (5, 6, 7):
                ws_det.cell(row=dr, column=col_l).alignment = center
                if not is_skipped_row:
                    ws_det.cell(row=dr, column=col_l).fill = blue_fill

            # Thick bottom border after last row of each developer
            is_last = bi == len(all_display_bundles) - 1
            _apply_borders(
                ws_det, dr, DET_COLS, thick_bottom if is_last else thin_border
            )

            # Wrap remarks and issues
            ws_det.cell(row=dr, column=16).alignment = wrap
            ws_det.cell(row=dr, column=17).alignment = wrap
            remarks_lines = len(bundle.reasoning) // 120 + 1
            ws_det.row_dimensions[dr].height = max(40, remarks_lines * 15)

    # ── Save ──────────────────────────────────────────────────────────────────
    output_dir = pathlib.Path("outputs") / team
    output_dir.mkdir(parents=True, exist_ok=True)
    filename = f"CodeQuality_Report_{team}_{year}_{month:02d}.xlsx"
    output_path = output_dir / filename
    wb.save(str(output_path))

    return str(output_path)
