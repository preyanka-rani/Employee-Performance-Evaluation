"""
app/services/data_sources/excel_parser.py
──────────────────────────────────────────
Parses uploaded TL Assessment Excel files into validated row objects.

Supported column sets
─────────────────────
Minimal (legacy upload endpoint):
  | employee_email | problem_solving | kpi | general |

Full (bulk-run endpoint):
  | employee_id | email | name | tl_problem_solving | tl_kpi | tl_general |
  | gitlab_username |  ← optional

Validation rules:
  - employee_email / email: required, valid email format
  - problem_solving / tl_problem_solving: 0 – 10
  - kpi / tl_kpi: 0 – 15
  - general / tl_general: 0 – 15

Returns a list of TLAssessmentRow (legacy) or TLRow (bulk-run).
Raises ExcelParseError with detailed messages on structural problems.
"""

import io
from dataclasses import dataclass, field
from typing import Any

import openpyxl
import pandas as pd
from pydantic import ValidationError

from app.core.logging_config import get_logger
from app.schemas.scores import TLAssessmentRow

logger = get_logger(__name__)


class ExcelParseError(Exception):
    """Raised when the uploaded Excel cannot be parsed or fails validation."""

    def __init__(self, errors: list[str]) -> None:
        self.errors = errors
        super().__init__("; ".join(errors))


# Mapping from human-friendly column names to normalised keys
_COLUMN_MAP: dict[str, str] = {
    "employee_email": "employee_email",
    "email": "employee_email",
    "problem_solving": "problem_solving",
    "problem solving": "problem_solving",
    "critical thinking": "problem_solving",
    "kpi": "kpi",
    "performance agreement": "kpi",
    "general": "general",
    "general assessment": "general",
    "team lead general assessment": "general",
}

_REQUIRED_KEYS: set[str] = {"employee_email", "problem_solving", "kpi", "general"}


@dataclass
class ExcelParseResult:
    rows: list[TLAssessmentRow]
    errors: list[str]


def parse_tl_assessment_excel(file_bytes: bytes) -> ExcelParseResult:
    """
    Parse a TL Assessment Excel upload.

    Args:
        file_bytes: Raw bytes of the uploaded .xlsx / .xls file.

    Returns:
        ExcelParseResult with valid rows and any per-row error messages.
    """
    errors: list[str] = []
    rows: list[TLAssessmentRow] = []

    try:
        workbook = openpyxl.load_workbook(
            io.BytesIO(file_bytes), read_only=True, data_only=True
        )
    except Exception as exc:
        raise ExcelParseError([f"Cannot open Excel file: {exc}"]) from exc

    sheet = workbook.active
    if sheet is None:
        raise ExcelParseError(["Excel workbook has no active sheet."])

    # ── Read header row ───────────────────────────────────────────────────────
    header_row = next(sheet.iter_rows(min_row=1, max_row=1, values_only=True), None)
    if header_row is None:
        raise ExcelParseError(["Excel file appears to be empty."])

    # Normalise header names
    col_index: dict[str, int] = {}
    for idx, cell_value in enumerate(header_row):
        if cell_value is None:
            continue
        normalised = _COLUMN_MAP.get(str(cell_value).strip().lower())
        if normalised:
            col_index[normalised] = idx

    missing_cols = _REQUIRED_KEYS - col_index.keys()
    if missing_cols:
        raise ExcelParseError(
            [f"Missing required columns: {', '.join(sorted(missing_cols))}"]
        )

    # ── Parse data rows ───────────────────────────────────────────────────────
    for row_num, row in enumerate(
        sheet.iter_rows(min_row=2, values_only=True), start=2
    ):
        if all(cell is None for cell in row):
            continue  # skip entirely blank rows

        def _get(key: str) -> str:
            val = row[col_index[key]]
            return str(val).strip() if val is not None else ""

        raw = {
            "employee_email": _get("employee_email").lower(),
            "problem_solving": _get("problem_solving"),
            "kpi": _get("kpi"),
            "general": _get("general"),
        }

        try:
            validated = TLAssessmentRow(**raw)
            rows.append(validated)
        except ValidationError as exc:
            for err in exc.errors():
                field_name = ".".join(str(loc) for loc in err["loc"])
                errors.append(f"Row {row_num}, field '{field_name}': {err['msg']}")
        except Exception as exc:
            errors.append(f"Row {row_num}: unexpected error – {exc}")

    workbook.close()
    logger.info("excel_parsed", valid_rows=len(rows), error_count=len(errors))
    return ExcelParseResult(rows=rows, errors=errors)


# ── Full-row parser for the bulk-run endpoint ─────────────────────────────────

# Canonical column name → accepted aliases (lowercase, underscored)
# Handles actual Excel headers like:
#   "critical_thinking_and_problem_solving_(10%)"
#   "annual__performance_agreement_(apa)_(15%)"
#   "team_leader_assessment_(15%)"
#   "employee_email"
_BULK_COLUMN_ALIASES: dict[str, list[str]] = {
    "employee_id": [
        "employee_id",
        "emp_id",
        "id",
        "sl",  # sometimes used as serial/ID
    ],
    "email": [
        "email",
        "employee_email",
        "email_address",
        "work_email",
    ],
    "name": [
        "name",
        "full_name",
        "employee_name",
        "employee",
        "resource_name",
    ],
    "tl_problem_solving": [
        "tl_problem_solving",
        "tl_ps",
        "ps",
        "problem_solving",
        "critical_thinking_and_problem_solving_(10%)",
        "critical_thinking_and_problem_solving_10",  # cleaned form
        "critical_thinking_&_problem_solving_(10%)",
        "critical_thinking_&_problem_solving_10",
        "critical_thinking_problem_solving_(10%)",
        "critical_thinking_problem_solving_10",
        "problem_solving_(10%)",
        "problem_solving_10",
        "critical_thinking",
    ],
    "tl_kpi": [
        "tl_kpi",
        "kpi",
        "annual__performance_agreement_(apa)_(15%)",
        "annual_performance_agreement_apa_15",  # cleaned form
        "annual_performance_agreement_(apa)_(15%)",
        "performance_agreement_(15%)",
        "performance_agreement_15",
        "performance_agreement",
        "apa_(15%)",
        "apa_15",
        "apa",
    ],
    "tl_general": [
        "tl_general",
        "general",
        "tl_general_assessment",
        "team_leader_assessment_(15%)",
        "team_leader_assessment_15",  # cleaned form
        "team_leader_general_assessment_(15%)",
        "team_leader_general_assessment_15",
        "team_lead_assessment_(15%)",
        "team_lead_assessment_15",
        "tl_assessment_(15%)",
        "tl_assessment_15",
        "general_assessment",
        "leadership_assessment",
    ],
    "gitlab_username": [
        "gitlab_username",
        "gitlab",
        "gitlab_user",
        "gitlab_id",
    ],
    "team": [
        "team",
        "team_name",
        "department",
    ],
}
_BULK_REQUIRED = {
    "employee_id",
    "email",
    "name",
    "tl_problem_solving",
    "tl_kpi",
    "tl_general",
}


@dataclass
class TLRow:
    """One row from the bulk-run TL marks spreadsheet (full employee record)."""

    employee_id: str
    email: str
    name: str
    tl_problem_solving: float
    tl_kpi: float
    tl_general: float
    gitlab_username: str | None = field(default=None)

    @property
    def tl_total(self) -> float:
        return round(self.tl_problem_solving + self.tl_kpi + self.tl_general, 4)


def _clean_col(name: str) -> str:
    """
    Normalise a raw column header so it matches alias keys:
      - strip whitespace
      - lowercase
      - replace spaces + hyphens with underscores
      - remove double underscores that appear after stripping parens/percent
    """
    import re

    s = str(name).strip().lower()
    # Replace spaces/hyphens with underscore
    s = re.sub(r"[\s\-]+", "_", s)
    # Remove characters that appear in headers like "(10%)", "&", etc.
    # but KEEP the core text and underscores
    s = re.sub(r"[()%&]", "", s)
    # Collapse multiple underscores
    s = re.sub(r"_+", "_", s)
    s = s.strip("_")
    return s


def _normalise_bulk_columns(df: pd.DataFrame) -> pd.DataFrame:
    """Strip, lowercase, remove special chars from headers; apply alias map."""
    df.columns = [_clean_col(c) for c in df.columns]
    rename_map: dict[str, str] = {}
    for canonical, aliases in _BULK_COLUMN_ALIASES.items():
        for alias in aliases:
            cleaned_alias = _clean_col(alias)
            if cleaned_alias in df.columns and canonical not in df.columns:
                rename_map[cleaned_alias] = canonical
                break
    return df.rename(columns=rename_map)


def parse_tl_excel(content: bytes) -> list[TLRow]:
    """
    Parse the raw bytes of a bulk-run TL marks Excel file.

    Args:
        content: Raw bytes from UploadFile.read().

    Returns:
        List of TLRow objects — one per valid, non-empty spreadsheet row.

    Raises:
        ExcelParseError: If the file is unreadable or missing required columns.
    """
    try:
        df = pd.read_excel(io.BytesIO(content), dtype=str)
    except Exception as exc:
        raise ExcelParseError([f"Cannot read Excel file: {exc}"]) from exc

    df = _normalise_bulk_columns(df)

    missing = _BULK_REQUIRED - set(df.columns)
    if missing:
        raise ExcelParseError(
            [
                f"Excel file is missing required columns: {sorted(missing)}. "
                f"Found: {sorted(df.columns.tolist())}"
            ]
        )

    # Drop rows where all required fields are NaN (trailing blank rows)
    df = df.dropna(subset=list(_BULK_REQUIRED), how="all").reset_index(drop=True)

    rows: list[TLRow] = []
    parse_errors: list[str] = []

    for idx, row_series in df.iterrows():
        row: dict[str, Any] = row_series.to_dict()
        row_num = int(idx) + 2  # +2 → header row + 1-based display

        try:
            emp_id = str(row["employee_id"]).strip()
            email = str(row["email"]).strip().lower()
            name = str(row["name"]).strip()

            # employee_id may be empty — leave it blank; the endpoint will
            # resolve it via a MySQL lookup before processing starts.
            if emp_id in ("nan", "none"):
                emp_id = ""
            if not email or "@" not in email:
                parse_errors.append(f"Row {row_num}: invalid email '{email}'")
                continue
            if not name or name in ("nan", "none"):
                parse_errors.append(f"Row {row_num}: name is empty")
                continue

            ps = min(max(float(row.get("tl_problem_solving") or 0), 0.0), 10.0)
            kpi = min(max(float(row.get("tl_kpi") or 0), 0.0), 15.0)
            general = min(max(float(row.get("tl_general") or 0), 0.0), 15.0)

            raw_gl = row.get("gitlab_username", "")
            gitlab = (
                str(raw_gl).strip()
                if raw_gl and str(raw_gl) not in ("nan", "none", "")
                else None
            )

            rows.append(
                TLRow(
                    employee_id=emp_id,
                    email=email,
                    name=name,
                    tl_problem_solving=ps,
                    tl_kpi=kpi,
                    tl_general=general,
                    gitlab_username=gitlab,
                )
            )
        except (ValueError, TypeError) as exc:
            parse_errors.append(f"Row {row_num}: {exc}")

    if parse_errors and not rows:
        raise ExcelParseError(parse_errors)

    logger.info(
        "bulk_excel_parsed",
        valid_rows=len(rows),
        error_count=len(parse_errors),
    )
    return rows
