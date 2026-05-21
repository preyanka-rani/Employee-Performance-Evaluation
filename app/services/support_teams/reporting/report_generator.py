"""
app/services/support_teams/reporting/report_generator.py
──────────────────────────────────────────────────────────
Generates formatted Excel reports for Support Team evaluation runs.

Report columns (23 total):
    Employee ID | Name | Email |
    Total Log Hours | Log Hours Score (0-100) | Sentiment Score (0-100) |
    CRM Log Score (0-100) |
    Total Tickets | Avg Taken Days | Tickets Volume Score | Tickets Speed Score |
    Tickets Evaluation Score (0-100) |
    Monthly Functional Score (0-100) | Segment A Marks (0-30) |
    Attendance Score (0-100) | Attendance Marks (0-10) |
    Support Readiness (0-10) | KPI (0-15) | General (0-15) | TL Total (0-40) |
    Segment B Marks (0-50) | Base Total (0-80) | Final Score (0-100)

Color coding on Final Score column:
    ≥ 80 → green
    60-79 → yellow
    < 60  → red

Saved to: outputs/reports/Support_Final_Report_{team}_{year}_{month:02d}.xlsx
"""

from __future__ import annotations

import pathlib

import openpyxl
from openpyxl.styles import Alignment, Font, PatternFill
from sqlalchemy import select
from sqlalchemy.ext.asyncio import AsyncSession

from app.core.logging_config import get_logger
from app.models.support_scores import SupportFinalScore
from app.repositories.employee_repository import EmployeeRepository

logger = get_logger(__name__)

# Output directory (matches developer report convention)
_REPORTS_DIR = pathlib.Path("outputs") / "reports"


async def generate_support_excel_report(
    run_id: int,
    emails: list[str],
    team: str,
    year: int,
    month: int,
    db: AsyncSession,
) -> str:
    """
    Build a formatted Excel report for a completed support team bulk evaluation.

    Reads from support_final_scores for all employees in ``emails`` that
    belong to the specified evaluation run.

    Returns:
        Absolute file path to the saved .xlsx report.
    """
    # ── Fetch scores ──────────────────────────────────────────────────────────
    result = await db.execute(
        select(SupportFinalScore).where(
            SupportFinalScore.evaluation_run_id == run_id,
            SupportFinalScore.employee_email.in_(emails),
        )
    )
    scores: list[SupportFinalScore] = list(result.scalars().all())

    # ── Employee name lookup ──────────────────────────────────────────────────
    emp_repo = EmployeeRepository(db)
    email_to_name: dict[str, str] = {}
    email_to_id: dict[str, str] = {}
    for email in emails:
        emp = await emp_repo.get_by_email(email)
        if emp is not None:
            email_to_name[email] = emp.name or email
            email_to_id[email] = emp.employee_id or ""

    # ── Workbook ──────────────────────────────────────────────────────────────
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = f"{team.upper()} {year}-{month:02d}"

    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill("solid", fgColor="2E4057")
    header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    data_align = Alignment(horizontal="center", vertical="center")

    green_fill = PatternFill("solid", fgColor="C6EFCE")
    yellow_fill = PatternFill("solid", fgColor="FFEB9C")
    red_fill = PatternFill("solid", fgColor="FFC7CE")

    headers = [
        "Employee ID",
        "Name",
        "Email",
        # CRM Log
        "Total Log Hours",
        "Log Hours Score (0-100)",
        "Sentiment Score (0-100)",
        "CRM Log Score (0-100)",
        # Tickets
        "Total Tickets",
        "Avg Resolution Days",
        "Tickets Volume Score (0-100)",
        "Tickets Speed Score (0-100)",
        "Tickets Evaluation Score (0-100)",
        # Segment A
        "Monthly Functional Score (0-100)",
        "Segment A Marks (0-30)",
        # Segment B
        "Attendance Score (0-100)",
        "Attendance Marks (0-10)",
        "Support Readiness (0-10)",
        "KPI Agreement (0-15)",
        "Leadership General (0-15)",
        "TL Total (0-40)",
        "Segment B Marks (0-50)",
        # Final
        "Base Total (0-80)",
        "Final Score (0-100)",
    ]

    ws.append(headers)

    # Style header row
    for col_idx in range(1, len(headers) + 1):
        cell = ws.cell(row=1, column=col_idx)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align

    ws.freeze_panes = "A2"
    ws.row_dimensions[1].height = 45

    # ── Data rows ─────────────────────────────────────────────────────────────
    # Sort descending by final score
    scores_sorted = sorted(scores, key=lambda s: s.final_score, reverse=True)

    for s in scores_sorted:
        emp_email = s.employee_email

        # Compute ticket volume/speed scores from stored raw data for display
        # (we read back the stored aggregates since tickets_evaluation_score is stored)
        from app.services.support_teams.scoring.formulas import (
            compute_monthly_tickets_score,
            compute_ticket_resolution_score,
        )

        vol_score = compute_monthly_tickets_score(s.total_tickets)
        spd_score = compute_ticket_resolution_score(s.average_taken_days)

        row_data = [
            email_to_id.get(emp_email, ""),
            email_to_name.get(emp_email, emp_email),
            emp_email,
            # CRM Log
            round(s.total_log_hours, 2),
            round(s.log_hours_score, 2),
            round(s.sentiment_score, 2),
            round(s.crm_log_score, 2),
            # Tickets
            s.total_tickets,
            round(s.average_taken_days, 2),
            round(vol_score, 2),
            round(spd_score, 2),
            round(s.tickets_evaluation_score, 2),
            # Segment A
            round(s.monthly_functional_score, 2),
            round(s.segment_a_marks, 2),
            # Segment B
            round(s.attendance_score, 2),
            round(s.attendance_marks, 2),
            round(s.support_readiness, 2),
            round(s.kpi, 2),
            round(s.general, 2),
            round(s.tl_total, 2),
            round(s.segment_b_marks, 2),
            # Final
            round(s.base_total, 2),
            round(s.final_score, 2),
        ]
        ws.append(row_data)

        # Align all data cells
        row_num = ws.max_row
        for col_idx in range(1, len(headers) + 1):
            ws.cell(row=row_num, column=col_idx).alignment = data_align

        # Color-code Final Score (last column)
        final_cell = ws.cell(row=row_num, column=len(headers))
        if s.final_score >= 80.0:
            final_cell.fill = green_fill
        elif s.final_score >= 60.0:
            final_cell.fill = yellow_fill
        else:
            final_cell.fill = red_fill

    # ── Auto-width ────────────────────────────────────────────────────────────
    for col_idx, header in enumerate(headers, start=1):
        col_letter = openpyxl.utils.get_column_letter(col_idx)
        # Estimate width from content
        max_width = max(
            len(header),
            max(
                (
                    len(str(ws.cell(row=r, column=col_idx).value or ""))
                    for r in range(2, ws.max_row + 1)
                ),
                default=10,
            ),
        )
        ws.column_dimensions[col_letter].width = min(max_width + 4, 30)

    # ── Add summary row ───────────────────────────────────────────────────────
    if scores_sorted:
        ws.append([])  # blank separator
        avg_final = round(
            sum(s.final_score for s in scores_sorted) / len(scores_sorted), 2
        )
        summary_row = [""] * len(headers)
        summary_row[0] = "AVERAGE"
        summary_row[-1] = avg_final
        ws.append(summary_row)
        avg_cell = ws.cell(row=ws.max_row, column=len(headers))
        avg_cell.font = Font(bold=True)
        if avg_final >= 80.0:
            avg_cell.fill = green_fill
        elif avg_final >= 60.0:
            avg_cell.fill = yellow_fill
        else:
            avg_cell.fill = red_fill

    # ── Save ──────────────────────────────────────────────────────────────────
    _REPORTS_DIR.mkdir(parents=True, exist_ok=True)
    filename = f"Support_Final_Report_{team}_{year}_{month:02d}.xlsx"
    output_path = _REPORTS_DIR / filename
    wb.save(str(output_path))

    logger.info(
        "support_report_saved",
        path=str(output_path),
        employee_count=len(scores_sorted),
        team=team,
        year=year,
        month=month,
    )

    return str(output_path.resolve())
