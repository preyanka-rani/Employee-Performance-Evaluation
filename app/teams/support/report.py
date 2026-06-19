"""
app/teams/support/report.py
───────────────────────────
Generates formatted Excel reports for Support Team evaluation runs.

Sheet 1 – "Detailed Report"
    Identity: Employee ID | Name | Email | <team col> | year | month_id
    Segment A: Total Log Hours | Log Hours Score | Sentiment Score | CRM Log Score |
               Total Tickets | Avg Resolution Days | Tickets Volume Score |
               Tickets Speed Score | Tickets Evaluation Score |
               Monthly Functional Score | Segment A Marks (0-30)
    Segment B: Attendance Score | Attendance Marks (0-10) |
               <TL readiness col> | <TL kpi col> | <TL general col> |
               TL Total (0-40) | Segment B Marks (0-50)
    Final:     Base Total (0-80) | Financial Contribution (0-20) | Total Score (0-100)

    Column names for team col and TL cols come from the original input Excel
    headers (passed via col_names dict).

Sheet 2 – "Final Summary" (matches EBS of Business Automation Ltd. format)
    Team | Employee | Email |
    avg_functional_score |
    <TL readiness col - original name> |
    avg_office_discipline_score |
    <TL kpi col - original name> |
    <TL general col - original name> |
    Avg_eval_scores | Percentage | Avg_eva_grade

Grade (applied to Percentage × 100):
    88–100 → A  |  84–87 → B  |  75–83 → C  |  0–74 → D

Saved to: outputs/support/Support_Final_Report_{team}_{year}_{month:02d}.xlsx
"""

from __future__ import annotations

import pathlib

import openpyxl
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from sqlalchemy import select
from sqlalchemy.ext.asyncio import AsyncSession

from app.core.logging_config import get_logger
from app.models.support_scores import SupportFinalScore
from app.repositories.employee_repository import EmployeeRepository
from app.teams.support.formulas import (
    compute_monthly_tickets_score,
    compute_ticket_resolution_score,
)

logger = get_logger(__name__)

# Output directory – support team reports live here
_REPORTS_DIR = pathlib.Path("outputs") / "support"

# ── Fallback column names (used when original headers are not available) ──────
_DEFAULT_COL_NAMES: dict[str, str] = {
    "team_name": "team_name",
    "support_readiness": "Support Readiness & Issue Handling (0-10)",
    "kpi": "KPI Agreement (0-15)",
    "general": "Leadership General Assessment (0-15)",
}


# ── Grade helper ──────────────────────────────────────────────────────────────


def _get_grade_from_pct(pct: float) -> str:
    """
    Return letter grade from a percentage value (0-100 scale).
    Thresholds match the EBS of Business Automation Ltd. format.
    """
    if pct >= 88:
        return "A"
    if pct >= 84:
        return "B"
    if pct >= 75:
        return "C"
    return "D"


async def generate_support_excel_report(
    run_id: int,
    emails: list[str],
    team: str,
    year: int,
    month: int,
    db: AsyncSession,
    col_names: dict[str, str] | None = None,
    team_display_name: str = "",
) -> str:
    """
    Build a formatted two-sheet Excel report for a completed support team
    bulk evaluation run.

    Args:
        col_names: canonical→original_header mapping from the uploaded Excel.
                   Used to preserve original column names in both sheets.
        team_display_name: raw team name value from the uploaded Excel (e.g.
                           "Tech Support").  Falls back to ``team`` key.

    Sheet 1: Full detailed breakdown for every employee.
    Sheet 2: EBS-format final summary (matches EBS of Business Automation Ltd.)

    Returns:
        Absolute file path to the saved .xlsx report.
    """
    # ── Resolve dynamic column names ──────────────────────────────────────────
    _cn = col_names or {}
    team_col_hdr = _cn.get("team_name", _DEFAULT_COL_NAMES["team_name"])
    tl_readiness_hdr = _cn.get(
        "support_readiness", _DEFAULT_COL_NAMES["support_readiness"]
    )
    tl_kpi_hdr = _cn.get("kpi", _DEFAULT_COL_NAMES["kpi"])
    tl_general_hdr = _cn.get("general", _DEFAULT_COL_NAMES["general"])
    team_display = team_display_name or team

    # ── Fetch scores ──────────────────────────────────────────────────────────
    result = await db.execute(
        select(SupportFinalScore).where(
            SupportFinalScore.evaluation_run_id == run_id,
            SupportFinalScore.employee_email.in_(emails),
        )
    )
    scores: list[SupportFinalScore] = list(result.scalars().all())

    # ── Employee lookup ───────────────────────────────────────────────────────
    emp_repo = EmployeeRepository(db)
    email_to_name: dict[str, str] = {}
    email_to_id: dict[str, str] = {}
    for email in emails:
        emp = await emp_repo.get_by_email(email)
        if emp is not None:
            email_to_name[email] = emp.name or email
            email_to_id[email] = emp.employee_id or ""

    # Sort descending by base_total for consistent ordering
    scores_sorted = sorted(scores, key=lambda s: s.base_total, reverse=True)

    # ── Common fill styles ────────────────────────────────────────────────────
    hdr_font = Font(bold=True, color="FFFFFF")
    hdr_fill = PatternFill("solid", fgColor="2E4057")
    hdr_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    center_align = Alignment(horizontal="center", vertical="center")

    green_fill = PatternFill("solid", fgColor="C6EFCE")  # ≥ 80
    yellow_fill = PatternFill("solid", fgColor="FFEB9C")  # 60-79
    red_fill = PatternFill("solid", fgColor="FFC7CE")  # < 60
    grade_a_fill = PatternFill("solid", fgColor="00B050")  # A – dark green
    grade_b_fill = PatternFill("solid", fgColor="92D050")  # B – light green
    grade_c_fill = PatternFill("solid", fgColor="FFEB9C")  # C – yellow
    grade_d_fill = PatternFill("solid", fgColor="FFC7CE")  # D – red
    section_fill = PatternFill("solid", fgColor="D6E4F0")  # section header rows
    total_fill = PatternFill("solid", fgColor="1F4E79")

    thin_side = Side(style="thin", color="AAAAAA")
    thin_border = Border(
        left=thin_side, right=thin_side, top=thin_side, bottom=thin_side
    )

    # ═════════════════════════════════════════════════════════════════════════
    # SHEET 1 – Detailed Report
    # ═════════════════════════════════════════════════════════════════════════
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Detailed Report"

    detail_headers = [
        "Employee ID",
        "Name",
        "Email",
        team_col_hdr,  # original header from input Excel (or "team_name")
        "year",
        "month_id",
        # ── Segment A: CRM Log ──
        "Total Log Hours",
        "Log Hours Score (0-100)",
        "Sentiment Score (0-100)",
        "CRM Log Score (0-100)",
        # ── Segment A: Tickets ──
        "Total Tickets",
        "Avg Resolution Days",
        "Tickets Volume Score (0-100)",
        "Tickets Speed Score (0-100)",
        "Tickets Evaluation Score (0-100)",
        # ── Segment A: Totals ──
        "Monthly Functional Score (0-100)",
        "Segment A Marks (0-30)",
        # ── Segment B ──
        "Attendance Score (0-100)",
        "Attendance Marks (0-10)",
        tl_readiness_hdr,  # original header from input Excel
        tl_kpi_hdr,  # original header from input Excel
        tl_general_hdr,  # original header from input Excel
        "TL Total (0-40)",
        "Segment B Marks (0-50)",
        # ── Final ──
        "Base Total (0-80)",
        "Financial Contribution (0-20)",
        "Total Score (0-100)",
    ]

    # Dynamic column widths (match order of detail_headers above)
    detail_widths = [
        14,
        22,
        30,
        max(18, len(team_col_hdr) + 2),  # team col
        8,
        10,
        18,
        22,
        22,
        20,  # Seg A CRM
        14,
        22,
        24,
        22,
        26,  # Seg A Tickets
        26,
        22,  # Seg A Totals
        22,
        20,  # Attendance
        max(24, len(tl_readiness_hdr) + 2),  # TL readiness
        max(20, len(tl_kpi_hdr) + 2),  # TL kpi
        max(24, len(tl_general_hdr) + 2),  # TL general
        16,
        22,
        18,
        24,
        18,  # TL Total, SegB, Base, FC, Total
    ]

    ws.append(detail_headers)
    ws.row_dimensions[1].height = 50
    for col_idx in range(1, len(detail_headers) + 1):
        cell = ws.cell(row=1, column=col_idx)
        cell.font = hdr_font
        cell.fill = hdr_fill
        cell.alignment = hdr_align
        cell.border = thin_border

    ws.freeze_panes = "A2"

    # Apply column widths
    for col_idx, width in enumerate(detail_widths, start=1):
        ws.column_dimensions[get_column_letter(col_idx)].width = width

    # Financial contribution placeholder (0 for everyone until feature is live)
    _FC = 0

    for s in scores_sorted:
        emp_email = s.employee_email
        vol_score = compute_monthly_tickets_score(s.total_tickets)
        spd_score = compute_ticket_resolution_score(s.average_taken_days)
        total_score = round(s.base_total + _FC, 2)

        row_data = [
            email_to_id.get(emp_email, ""),
            email_to_name.get(emp_email, emp_email),
            emp_email,
            team_display,  # original team name from the uploaded Excel
            year,
            month,
            # Segment A – CRM Log
            round(s.total_log_hours, 2),
            round(s.log_hours_score, 2),
            round(s.sentiment_score, 2),
            round(s.crm_log_score, 2),
            # Segment A – Tickets
            s.total_tickets,
            round(s.average_taken_days, 2),
            round(vol_score, 2),
            round(spd_score, 2),
            round(s.tickets_evaluation_score, 2),
            # Segment A – Totals
            round(s.monthly_functional_score, 2),
            round(s.segment_a_marks, 2),
            # Segment B
            round(s.attendance_score, 2),
            round(s.attendance_marks, 2),
            round(s.support_readiness, 2),
            round(s.kpi, 2),
            round(s.general, 2),
            round(s.tl_total, 2),
            round(s.segment_b_marks, 2),
            # Final
            round(s.base_total, 2),
            _FC,
            total_score,
        ]
        ws.append(row_data)

        row_num = ws.max_row
        for col_idx in range(1, len(detail_headers) + 1):
            c = ws.cell(row=row_num, column=col_idx)
            c.alignment = center_align
            c.border = thin_border

        # Color-code Total Score (last column)
        total_cell = ws.cell(row=row_num, column=len(detail_headers))
        if total_score >= 80.0:
            total_cell.fill = green_fill
        elif total_score >= 60.0:
            total_cell.fill = yellow_fill
        else:
            total_cell.fill = red_fill

    # Summary / average row
    if scores_sorted:
        ws.append([])
        totals = [round(s.base_total + _FC, 2) for s in scores_sorted]
        avg_total = round(sum(totals) / len(totals), 2)
        summary = [""] * len(detail_headers)
        summary[0] = "TEAM AVERAGE"
        summary[-3] = round(
            sum(s.base_total for s in scores_sorted) / len(scores_sorted), 2
        )
        summary[-2] = _FC
        summary[-1] = avg_total
        ws.append(summary)
        last_row = ws.max_row
        for col_idx in range(1, len(detail_headers) + 1):
            c = ws.cell(row=last_row, column=col_idx)
            c.font = Font(bold=True)
            c.border = thin_border
        avg_cell = ws.cell(row=last_row, column=len(detail_headers))
        avg_cell.fill = (
            green_fill
            if avg_total >= 80
            else yellow_fill if avg_total >= 60 else red_fill
        )

    # ═════════════════════════════════════════════════════════════════════════
    # SHEET 2 – Final Summary (EBS of Business Automation Ltd. format)
    #
    # Columns (11):
    #   Team | Employee | Email |
    #   avg_functional_score |
    #   <TL readiness - original name> |
    #   avg_office_discipline_score |
    #   <TL kpi - original name> |
    #   <TL general - original name> |
    #   Avg_eval_scores | Percentage | Avg_eva_grade
    # ═════════════════════════════════════════════════════════════════════════

    ws2 = wb.create_sheet("Final Summary")

    summary_headers = [
        "Team",
        "Employee",
        "Email",
        "avg_functional_score",
        tl_readiness_hdr,  # original name from input Excel
        "avg_office_discipline_score",
        tl_kpi_hdr,  # original name from input Excel
        tl_general_hdr,  # original name from input Excel
        "Avg_eval_scores",
        "Percentage",
        "Avg_eva_grade",
    ]

    ws2.append(summary_headers)
    ws2.row_dimensions[1].height = 50
    for col_idx in range(1, len(summary_headers) + 1):
        cell = ws2.cell(row=1, column=col_idx)
        cell.font = hdr_font
        cell.fill = hdr_fill
        cell.alignment = hdr_align
        cell.border = thin_border

    ws2.freeze_panes = "A2"

    # Column widths
    summary_col_widths = [
        max(18, len(team_display) + 2),  # Team
        24,  # Employee
        32,  # Email
        20,  # avg_functional_score
        max(24, len(tl_readiness_hdr) + 2),  # TL readiness (original name)
        26,  # avg_office_discipline_score
        max(20, len(tl_kpi_hdr) + 2),  # TL kpi (original name)
        max(24, len(tl_general_hdr) + 2),  # TL general (original name)
        18,  # Avg_eval_scores
        12,  # Percentage
        14,  # Avg_eva_grade
    ]
    for col_idx, width in enumerate(summary_col_widths, start=1):
        ws2.column_dimensions[get_column_letter(col_idx)].width = width

    grade_fills = {
        "A": grade_a_fill,
        "B": grade_b_fill,
        "C": grade_c_fill,
        "D": grade_d_fill,
    }

    _MAX_SCORE = 80.0  # maximum possible Avg_eval_scores

    for s in scores_sorted:
        emp_email = s.employee_email
        avg_eval = round(
            s.segment_a_marks
            + s.support_readiness
            + s.attendance_marks
            + s.kpi
            + s.general,
            2,
        )
        percentage = round(avg_eval / _MAX_SCORE, 2)
        grade = _get_grade_from_pct(round(percentage * 100, 2))

        ws2.append(
            [
                team_display,
                email_to_name.get(emp_email, emp_email),
                emp_email,
                round(s.segment_a_marks, 2),  # avg_functional_score
                round(s.support_readiness, 2),  # TL readiness (original name)
                round(s.attendance_marks, 2),  # avg_office_discipline_score
                round(s.kpi, 2),  # TL kpi (original name)
                round(s.general, 2),  # TL general (original name)
                avg_eval,  # Avg_eval_scores
                percentage,  # Percentage (decimal, e.g. 0.82)
                grade,  # Avg_eva_grade
            ]
        )

        row_num = ws2.max_row
        for col_idx in range(1, len(summary_headers) + 1):
            c = ws2.cell(row=row_num, column=col_idx)
            c.alignment = center_align
            c.border = thin_border

        # Colour-code grade cell (column 11)
        grade_cell = ws2.cell(row=row_num, column=11)
        grade_cell.font = Font(
            bold=True, color="FFFFFF" if grade in ("A", "B") else "000000"
        )
        grade_cell.fill = grade_fills[grade]

        # Colour Percentage cell (column 10)
        pct_cell = ws2.cell(row=row_num, column=10)
        pct_val = percentage * 100
        pct_cell.fill = (
            grade_a_fill
            if pct_val >= 88
            else (
                grade_b_fill
                if pct_val >= 84
                else grade_c_fill if pct_val >= 75 else grade_d_fill
            )
        )

    # Team-average footer row
    if scores_sorted:
        ws2.append([])  # blank separator
        avg_seg_a = round(
            sum(s.segment_a_marks for s in scores_sorted) / len(scores_sorted), 2
        )
        avg_readiness = round(
            sum(s.support_readiness for s in scores_sorted) / len(scores_sorted), 2
        )
        avg_attendance = round(
            sum(s.attendance_marks for s in scores_sorted) / len(scores_sorted), 2
        )
        avg_kpi = round(sum(s.kpi for s in scores_sorted) / len(scores_sorted), 2)
        avg_general = round(
            sum(s.general for s in scores_sorted) / len(scores_sorted), 2
        )
        avg_eval_all = round(
            avg_seg_a + avg_readiness + avg_attendance + avg_kpi + avg_general, 2
        )
        avg_pct = round(avg_eval_all / _MAX_SCORE, 2)
        avg_grade = _get_grade_from_pct(round(avg_pct * 100, 2))

        ws2.append(
            [
                "TEAM AVERAGE",
                "",
                "",
                avg_seg_a,
                avg_readiness,
                avg_attendance,
                avg_kpi,
                avg_general,
                avg_eval_all,
                avg_pct,
                avg_grade,
            ]
        )
        last_row = ws2.max_row
        for col_idx in range(1, len(summary_headers) + 1):
            c = ws2.cell(row=last_row, column=col_idx)
            c.font = Font(bold=True, color="FFFFFF" if col_idx == 1 else "000000")
            c.fill = (
                PatternFill("solid", fgColor="1F4E79") if col_idx == 1 else section_fill
            )
            c.border = thin_border
            c.alignment = center_align

    # ── Save ──────────────────────────────────────────────────────────────────
    _REPORTS_DIR.mkdir(parents=True, exist_ok=True)
    filename = f"Support_Final_Report_{team}_{year}_{month:02d}.xlsx"
    output_path = _REPORTS_DIR / filename
    wb.save(str(output_path))

    logger.info(
        "support_report_saved",
        path=str(output_path),
        employee_count=len(scores_sorted),
        team=team,
        year=year,
        month=month,
    )

    return str(output_path.resolve())
