from __future__ import annotations

import pathlib
from typing import Any

import openpyxl
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from sqlalchemy import select
from sqlalchemy.ext.asyncio import AsyncSession

from app.core.logging_config import get_logger
from app.models.scores import (
    AttendanceScore,
    FinalScore,
    SentimentScore,
    WorkLogScore,
)
from app.repositories.employee_repository import EmployeeRepository

logger = get_logger(__name__)

_REPORTS_DIR = pathlib.Path("outputs") / "supply_chain"

_DEFAULT_COL_NAMES: dict[str, str] = {
    "team_name": "team_name",
    "problem_solving": "Critical Thinking & Problem Solving (0-10)",
    "kpi": "KPI Agreement (0-15)",
    "general": "General Assessment (0-15)",
}


def _get_grade_from_pct(pct: float) -> str:
    if pct >= 88:
        return "A"
    if pct >= 84:
        return "B"
    if pct >= 75:
        return "C"
    return "D"


async def _load_score_rows(
    db: AsyncSession,
    run_id: int,
    emails: list[str],
) -> dict[str, dict[str, Any]]:
    if not emails:
        return {}

    wl_rows = (
        (await db.execute(
            select(WorkLogScore).where(
                WorkLogScore.evaluation_run_id == run_id,
                WorkLogScore.employee_email.in_(emails),
            )
        )).scalars().all()
    )

    sent_rows = (
        (await db.execute(
            select(SentimentScore).where(
                SentimentScore.evaluation_run_id == run_id,
                SentimentScore.employee_email.in_(emails),
            )
        )).scalars().all()
    )

    att_rows = (
        (await db.execute(
            select(AttendanceScore).where(
                AttendanceScore.evaluation_run_id == run_id,
                AttendanceScore.employee_email.in_(emails),
            )
        )).scalars().all()
    )

    fs_rows = (
        (await db.execute(
            select(FinalScore).where(
                FinalScore.evaluation_run_id == run_id,
                FinalScore.employee_email.in_(emails),
            )
        )).scalars().all()
    )

    by_email: dict[str, dict[str, Any]] = {}
    for fs in fs_rows:
        by_email[fs.employee_email] = {
            "total_log_hours": 0.0,
            "log_hours_score": 0.0,
            "sentiment_score": 0.0,
            "avg_polarity": 0.0,
            "total_logs_analyzed": 0,
            "attendance_score": float(fs.attendance_score or 0.0),
            "attendance_marks": float(fs.attendance_marks or 0.0),
            "problem_solving": float(fs.problem_solving or 0.0),
            "kpi": float(fs.kpi or 0.0),
            "general": float(fs.general_assessment or 0.0),
            "tl_total": float(fs.tl_total or 0.0),
            "segment_a_marks": float(fs.segment_a_marks or 0.0),
            "segment_b_marks": float(fs.segment_b_marks or 0.0),
            "base_total": float(fs.base_total or 0.0),
            "final_score": float(fs.final_score or 0.0),
        }

    for wl in wl_rows:
        rec = by_email.get(wl.employee_email)
        if rec is not None:
            rec["total_log_hours"] = float(wl.total_log_hours or 0.0)
            rec["log_hours_score"] = float(wl.normalized_score or 0.0)

    for s in sent_rows:
        rec = by_email.get(s.employee_email)
        if rec is not None:
            rec["sentiment_score"] = float(s.score or 0.0)
            rec["avg_polarity"] = float(s.average_polarity or 0.0)
            rec["total_logs_analyzed"] = int(s.total_logs_analyzed or 0)

    for a in att_rows:
        rec = by_email.get(a.employee_email)
        if rec is not None:
            rec.setdefault("present_days", int(a.present_days or 0))
            rec.setdefault("late_attendance", int(a.late_attendance or 0))

    return by_email


async def generate_supply_chain_excel_report(
    run_id: int,
    emails: list[str],
    team: str,
    year: int,
    month: int,
    db: AsyncSession,
    col_names: dict[str, str] | None = None,
    team_display_name: str = "",
) -> str:
    _cn = col_names or {}
    team_col_hdr = _cn.get("team_name", _DEFAULT_COL_NAMES["team_name"])
    tl_ps_hdr = _cn.get("problem_solving", _DEFAULT_COL_NAMES["problem_solving"])
    tl_kpi_hdr = _cn.get("kpi", _DEFAULT_COL_NAMES["kpi"])
    tl_general_hdr = _cn.get("general", _DEFAULT_COL_NAMES["general"])
    team_display = team_display_name or team

    by_email = await _load_score_rows(db, run_id=run_id, emails=emails)

    emp_repo = EmployeeRepository(db)
    email_to_name: dict[str, str] = {}
    email_to_id: dict[str, str] = {}
    for email in emails:
        emp = await emp_repo.get_by_email(email)
        if emp is not None:
            email_to_name[email] = emp.name or email
            email_to_id[email] = emp.employee_id or ""

    scores_sorted = sorted(
        by_email.values(),
        key=lambda r: r["final_score"],
        reverse=True,
    )

    hdr_font = Font(bold=True, color="FFFFFF")
    hdr_fill = PatternFill("solid", fgColor="1F4E78")
    hdr_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    center_align = Alignment(horizontal="center", vertical="center")

    green_fill = PatternFill("solid", fgColor="C6EFCE")
    yellow_fill = PatternFill("solid", fgColor="FFEB9C")
    red_fill = PatternFill("solid", fgColor="FFC7CE")
    grade_a_fill = PatternFill("solid", fgColor="00B050")
    grade_b_fill = PatternFill("solid", fgColor="92D050")
    grade_c_fill = PatternFill("solid", fgColor="FFEB9C")
    grade_d_fill = PatternFill("solid", fgColor="FFC7CE")
    section_fill = PatternFill("solid", fgColor="D6E4F0")

    thin_side = Side(style="thin", color="AAAAAA")
    thin_border = Border(left=thin_side, right=thin_side, top=thin_side, bottom=thin_side)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Detailed Report"

    detail_headers = [
        "Employee ID", "Name", "Email", team_col_hdr, "year", "month",
        "Total Log Hours", "Log Hours Score (0-100)", "Sentiment Score (0-100)",
        "CRM Log Score (0-100)", "Monthly Functional Score (0-100)",
        "Segment A Marks (0-30)",
        "Attendance Score (0-100)", "Attendance Marks (0-10)",
        tl_ps_hdr, tl_kpi_hdr, tl_general_hdr,
        "TL Total (0-40)", "Segment B Marks (0-50)",
        "Base Total (0-80)", "Final Score (0-100)",
    ]

    detail_widths = [
        14, 22, 30, max(18, len(team_col_hdr) + 2), 8, 8,
        18, 22, 22, 22, 28, 22,
        22, 20, max(28, len(tl_ps_hdr) + 2), max(20, len(tl_kpi_hdr) + 2),
        max(24, len(tl_general_hdr) + 2), 16, 22, 18, 18,
    ]

    ws.append(detail_headers)
    ws.row_dimensions[1].height = 50
    for col_idx in range(1, len(detail_headers) + 1):
        cell = ws.cell(row=1, column=col_idx)
        cell.font = hdr_font
        cell.fill = hdr_fill
        cell.alignment = hdr_align
        cell.border = thin_border

    ws.freeze_panes = "A2"

    for col_idx, width in enumerate(detail_widths, start=1):
        ws.column_dimensions[get_column_letter(col_idx)].width = width

    for rec in scores_sorted:
        emp_email = next((e for e, r in by_email.items() if r is rec), "")
        if not emp_email:
            for e, r in by_email.items():
                if r is rec:
                    emp_email = e
                    break

        row_data = [
            email_to_id.get(emp_email, ""),
            email_to_name.get(emp_email, emp_email),
            emp_email,
            team_display,
            year, month,
            round(rec["total_log_hours"], 2),
            round(rec["log_hours_score"], 2),
            round(rec["sentiment_score"], 2),
            round(rec["log_hours_score"] * 0.9 + rec["sentiment_score"] * 0.1, 2),
            round(rec["segment_a_marks"] / 0.30 if rec["segment_a_marks"] else 0.0, 2),
            round(rec["segment_a_marks"], 2),
            round(rec["attendance_score"], 2),
            round(rec["attendance_marks"], 2),
            round(rec["problem_solving"], 2),
            round(rec["kpi"], 2),
            round(rec["general"], 2),
            round(rec["tl_total"], 2),
            round(rec["segment_b_marks"], 2),
            round(rec["base_total"], 2),
            round(rec["final_score"], 2),
        ]
        ws.append(row_data)

        row_num = ws.max_row
        for col_idx in range(1, len(detail_headers) + 1):
            c = ws.cell(row=row_num, column=col_idx)
            c.alignment = center_align
            c.border = thin_border

        final_cell = ws.cell(row=row_num, column=len(detail_headers))
        final_score = rec["final_score"]
        if final_score >= 80.0:
            final_cell.fill = green_fill
        elif final_score >= 60.0:
            final_cell.fill = yellow_fill
        else:
            final_cell.fill = red_fill

    if scores_sorted:
        ws.append([])
        avg_final = round(sum(r["final_score"] for r in scores_sorted) / len(scores_sorted), 2)
        summary = [""] * len(detail_headers)
        summary[0] = "TEAM AVERAGE"
        avg_monthly_func = round(
            sum((r["segment_a_marks"] / 0.30) if r["segment_a_marks"] else 0.0 for r in scores_sorted)
            / len(scores_sorted), 2
        )
        avg_seg_a = round(sum(r["segment_a_marks"] for r in scores_sorted) / len(scores_sorted), 2)
        avg_seg_b = round(sum(r["segment_b_marks"] for r in scores_sorted) / len(scores_sorted), 2)
        avg_base = round(sum(r["base_total"] for r in scores_sorted) / len(scores_sorted), 2)
        summary[6] = round(sum(r["total_log_hours"] for r in scores_sorted) / len(scores_sorted), 2)
        summary[7] = round(sum(r["log_hours_score"] for r in scores_sorted) / len(scores_sorted), 2)
        summary[8] = round(sum(r["sentiment_score"] for r in scores_sorted) / len(scores_sorted), 2)
        summary[9] = round(sum(r["log_hours_score"] * 0.9 + r["sentiment_score"] * 0.1 for r in scores_sorted) / len(scores_sorted), 2)
        summary[10] = avg_monthly_func
        summary[11] = avg_seg_a
        summary[12] = round(sum(r["attendance_score"] for r in scores_sorted) / len(scores_sorted), 2)
        summary[13] = round(sum(r["attendance_marks"] for r in scores_sorted) / len(scores_sorted), 2)
        summary[14] = round(sum(r["problem_solving"] for r in scores_sorted) / len(scores_sorted), 2)
        summary[15] = round(sum(r["kpi"] for r in scores_sorted) / len(scores_sorted), 2)
        summary[16] = round(sum(r["general"] for r in scores_sorted) / len(scores_sorted), 2)
        summary[17] = round(sum(r["tl_total"] for r in scores_sorted) / len(scores_sorted), 2)
        summary[18] = avg_seg_b
        summary[19] = avg_base
        summary[20] = avg_final
        ws.append(summary)
        last_row = ws.max_row
        for col_idx in range(1, len(detail_headers) + 1):
            c = ws.cell(row=last_row, column=col_idx)
            c.font = Font(bold=True)
            c.border = thin_border
        avg_cell = ws.cell(row=last_row, column=len(detail_headers))
        avg_cell.fill = green_fill if avg_final >= 80 else yellow_fill if avg_final >= 60 else red_fill

    ws2 = wb.create_sheet("Final Summary")

    summary_headers = [
        "Team", "Employee", "Email",
        "avg_functional_score", tl_ps_hdr,
        "avg_office_discipline_score", tl_kpi_hdr, tl_general_hdr,
        "Avg_eval_scores", "Percentage", "Avg_eva_grade",
    ]

    ws2.append(summary_headers)
    ws2.row_dimensions[1].height = 50
    for col_idx in range(1, len(summary_headers) + 1):
        cell = ws2.cell(row=1, column=col_idx)
        cell.font = hdr_font
        cell.fill = hdr_fill
        cell.alignment = hdr_align
        cell.border = thin_border

    ws2.freeze_panes = "A2"

    summary_col_widths = [
        max(18, len(team_display) + 2), 24, 32, 20,
        max(28, len(tl_ps_hdr) + 2), 26,
        max(20, len(tl_kpi_hdr) + 2), max(24, len(tl_general_hdr) + 2),
        18, 12, 14,
    ]
    for col_idx, width in enumerate(summary_col_widths, start=1):
        ws2.column_dimensions[get_column_letter(col_idx)].width = width

    grade_fills = {"A": grade_a_fill, "B": grade_b_fill, "C": grade_c_fill, "D": grade_d_fill}
    max_score = 80.0

    for rec in scores_sorted:
        emp_email = next((e for e, r in by_email.items() if r is rec), "")
        if not emp_email:
            for e, r in by_email.items():
                if r is rec:
                    emp_email = e
                    break

        avg_eval = round(
            rec["segment_a_marks"] + rec["problem_solving"]
            + rec["attendance_marks"] + rec["kpi"] + rec["general"],
            2,
        )
        percentage = round(avg_eval / max_score, 2)
        grade = _get_grade_from_pct(round(percentage * 100, 2))

        ws2.append([
            team_display,
            email_to_name.get(emp_email, emp_email),
            emp_email,
            round(rec["segment_a_marks"], 2),
            round(rec["problem_solving"], 2),
            round(rec["attendance_marks"], 2),
            round(rec["kpi"], 2),
            round(rec["general"], 2),
            avg_eval,
            percentage,
            grade,
        ])

        row_num = ws2.max_row
        for col_idx in range(1, len(summary_headers) + 1):
            c = ws2.cell(row=row_num, column=col_idx)
            c.alignment = center_align
            c.border = thin_border

        grade_cell = ws2.cell(row=row_num, column=11)
        grade_cell.font = Font(bold=True, color="FFFFFF" if grade in ("A", "B") else "000000")
        grade_cell.fill = grade_fills[grade]

        pct_cell = ws2.cell(row=row_num, column=10)
        pct_val = percentage * 100
        pct_cell.fill = (
            grade_a_fill if pct_val >= 88 else
            grade_b_fill if pct_val >= 84 else
            grade_c_fill if pct_val >= 75 else
            grade_d_fill
        )

    if scores_sorted:
        ws2.append([])
        avg_seg_a = round(sum(r["segment_a_marks"] for r in scores_sorted) / len(scores_sorted), 2)
        avg_ps = round(sum(r["problem_solving"] for r in scores_sorted) / len(scores_sorted), 2)
        avg_attendance = round(sum(r["attendance_marks"] for r in scores_sorted) / len(scores_sorted), 2)
        avg_kpi = round(sum(r["kpi"] for r in scores_sorted) / len(scores_sorted), 2)
        avg_general = round(sum(r["general"] for r in scores_sorted) / len(scores_sorted), 2)
        avg_eval_all = round(avg_seg_a + avg_ps + avg_attendance + avg_kpi + avg_general, 2)
        avg_pct = round(avg_eval_all / max_score, 2)
        avg_grade = _get_grade_from_pct(round(avg_pct * 100, 2))

        ws2.append([
            "TEAM AVERAGE", "", "",
            avg_seg_a, avg_ps, avg_attendance, avg_kpi, avg_general,
            avg_eval_all, avg_pct, avg_grade,
        ])
        last_row = ws2.max_row
        for col_idx in range(1, len(summary_headers) + 1):
            c = ws2.cell(row=last_row, column=col_idx)
            c.font = Font(bold=True, color="FFFFFF" if col_idx == 1 else "000000")
            c.fill = PatternFill("solid", fgColor="1F4E79") if col_idx == 1 else section_fill
            c.border = thin_border
            c.alignment = center_align

    _REPORTS_DIR.mkdir(parents=True, exist_ok=True)
    filename = f"Supply_Chain_Final_Report_{team}_{year}_{month:02d}.xlsx"
    output_path = _REPORTS_DIR / filename
    wb.save(str(output_path))

    logger.info("supply_chain_report_saved", path=str(output_path), employee_count=len(scores_sorted))
    return str(output_path.resolve())
