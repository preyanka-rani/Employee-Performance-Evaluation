"""
app/teams/sqa/report.py
────────────────────────
SQA worker report generator — adapted from the developer report.

Output paths:
    outputs/sqa/CodeQuality_Report_sqa_{year}_{month:02d}.xlsx
    outputs/sqa/Final_Report_sqa_{year}_{month:02d}.xlsx

The Final Report includes a ``summary_grade`` sheet with:
  - Renamed columns per SQA naming conventions
  - Grade calculation (88-100=A, 84-87=B, 75-83=C, 0-74=D)
  - Color-coded grades and exact team name matching the input Excel.
"""

from __future__ import annotations

import pathlib

import openpyxl
import pandas as pd
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from sqlalchemy import select
from sqlalchemy.ext.asyncio import AsyncSession

from app.core.logging_config import get_logger
from app.models.employee import Employee
from app.repositories.score_repository import DeveloperFinalScoreRepository
from app.services.reporting import report_generator

logger = get_logger(__name__)


def _get_sqa_grade(pct: float) -> str:
    if pct >= 88:
        return "A"
    if pct >= 84:
        return "B"
    if pct >= 75:
        return "C"
    return "D"


async def generate_sqa_reports(
    *,
    run_id: int,
    emails: list[str],
    year: int,
    month: int,
    db: AsyncSession,
) -> dict[str, str]:
    log = logger.bind(run_id=run_id, year=year, month=month, count=len(emails))
    log.info("sqa_reports_start")

    cq_path = await report_generator.generate_code_quality_report(
        run_id=run_id,
        emails=emails,
        team="sqa",
        year=year,
        month=month,
        db=db,
    )

    fr_path = await _generate_sqa_final_report(
        run_id=run_id,
        emails=emails,
        year=year,
        month=month,
        db=db,
    )

    log.info("sqa_reports_done", code_quality=cq_path, final=fr_path)
    return {
        "code_quality_report": str(pathlib.Path(cq_path).resolve()) if cq_path else "",
        "final_report": str(pathlib.Path(fr_path).resolve()) if fr_path else "",
    }


async def _generate_sqa_final_report(
    run_id: int,
    emails: list[str],
    year: int,
    month: int,
    db: AsyncSession,
) -> str:
    log = logger.bind(run_id=run_id, year=year, month=month, count=len(emails))
    dev_repo = DeveloperFinalScoreRepository(db)
    scores = await dev_repo.get_by_run_id(run_id=run_id, emails=emails)

    # ── Fetch Exact Team Names from Database ──
    result = await db.execute(select(Employee).where(Employee.email.in_(emails)))
    employees = result.scalars().all()
    team_map = {emp.email: emp.team for emp in employees}

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = f"sqa {year}-{month:02d}"

    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill("solid", fgColor="1F4E79")
    header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    green_fill = PatternFill("solid", fgColor="C6EFCE")
    yellow_fill = PatternFill("solid", fgColor="FFEB9C")
    red_fill = PatternFill("solid", fgColor="FFC7CE")

    # ── Updated Headers: Added 'Team' Column ──
    headers = [
        "Employee ID",
        "Name",
        "Email",
        "Team",  
        "Component 1 Score (0-100)",
        "  Code Quality (30%)",
        "  Resolution Rate % (35%)",
        "  Reopen Quality (15%)",
        "  Lines Added Score (10%)",
        "  Lines Deleted Score (10%)",
        "Work Log Hours",
        "Work Log Score (0-100)",
        "Sentiment Score (0-100)",
        "Component 2 Score (0-100)",
        "Attendance Score (0-100)",
        "TL Problem Solving (0-10)",
        "TL KPI (0-15)",
        "TL General (0-15)",
        "TL Total (0-40)",
        "Segment A Marks (0-30)",
        "Segment B Marks (0-50)",
        "Base Total (0-80)",
        "Final Score",
    ]
    ws.append(headers)

    for col_idx in range(1, len(headers) + 1):
        cell = ws.cell(row=1, column=col_idx)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment

    ws.freeze_panes = "A2"

    for s in sorted(scores, key=lambda x: x.final_score, reverse=True):
        row_data = [
            s.employee_id,
            s.employee_name,
            s.employee_email,
            team_map.get(s.employee_email, "sqa"),  # Dynamic Team Name inserted here
            round(s.component1_score, 2),
            round(s.code_quality_score, 2),
            round(s.resolution_rate, 2),
            round(s.reopen_quality_score, 2),
            round(s.lines_added_score, 2),
            round(s.lines_deleted_score, 2),
            round(s.work_log_hours, 2),
            round(s.work_log_score, 2),
            round(s.sentiment_score, 2),
            round(s.component2_score, 2),
            round(s.attendance_score, 2),
            round(s.tl_problem_solving, 2),
            round(s.tl_kpi, 2),
            round(s.tl_general, 2),
            round(s.tl_total, 2),
            round(s.segment_a_marks, 2),
            round(s.segment_b_marks, 2),
            round(s.base_total, 2),
            round(s.final_score, 2),
        ]
        ws.append(row_data)

        final_cell = ws.cell(row=ws.max_row, column=len(headers))
        if s.final_score >= 80:
            final_cell.fill = green_fill
        elif s.final_score >= 60:
            final_cell.fill = yellow_fill
        else:
            final_cell.fill = red_fill

    # Adjusted column widths to account for the new Team column
    column_widths = [
        14, 22, 30, 20, 24, 22, 24, 22, 22, 22, 18, 22, 22, 24, 22, 24, 14, 18, 14, 22, 22, 20, 18
    ]
    for col_idx, width in enumerate(column_widths, start=1):
        ws.column_dimensions[get_column_letter(col_idx)].width = width

    if scores:
        avg_final = round(sum(s.final_score for s in scores) / len(scores), 2)
        ws.append([""] * (len(headers) - 2) + ["Team Average", avg_final])
        summary_row = ws.max_row
        for col_idx in range(1, len(headers) + 1):
            ws.cell(row=summary_row, column=col_idx).font = Font(bold=True)

    output_dir = pathlib.Path("outputs") / "sqa"
    output_dir.mkdir(parents=True, exist_ok=True)
    filename = f"Final_Report_sqa_{year}_{month:02d}.xlsx"
    output_path = output_dir / filename
    wb.save(str(output_path))

    # ── Append Formatting-rich summary_grade sheet ────────────────────────
    try:
        df = pd.read_excel(output_path, sheet_name=0)

        summary_df = pd.DataFrame()
        summary_df["emp_email"] = df.get("Email", "")
        summary_df["emp_name"] = df.get("Name", "")
        summary_df["team_name"] = df.get("Team", "")  # Fetches exact team name
        
        summary_df["avg_functional_job_performance_30"] = df.get("Segment A Marks (0-30)", 0.0)
        summary_df["avg_office_discipline_10"] = df.get("Attendance Score (0-100)", 0.0) / 10.0
        summary_df["avg_critical_thinking_and_problem_solving_10"] = df.get("TL Problem Solving (0-10)", 0.0)
        summary_df["avg_monthly_performance_agreement_15"] = df.get("TL KPI (0-15)", 0.0)
        summary_df["avg_team_leader_assessment_15"] = df.get("TL General (0-15)", 0.0)
        
        summary_df["avg_total_scores"] = (
            summary_df["avg_functional_job_performance_30"]
            + summary_df["avg_office_discipline_10"]
            + summary_df["avg_critical_thinking_and_problem_solving_10"]
            + summary_df["avg_monthly_performance_agreement_15"]
            + summary_df["avg_team_leader_assessment_15"]
        )
        
        summary_df["reword_score_5"] = 0.0  # Added to match exact developer column structure
        summary_df["finalize_score"] = df.get("Final Score", 0.0) 
        summary_df["score_percentage"] = (summary_df["finalize_score"] / 100.0).round(4)
        summary_df["Avg_eva_grade"] = summary_df["finalize_score"].apply(_get_sqa_grade)

        with pd.ExcelWriter(
            output_path, engine="openpyxl", mode="a", if_sheet_exists="replace"
        ) as writer:
            summary_df.to_excel(writer, sheet_name="summary_grade", index=False)

            workbook = writer.book
            worksheet = writer.sheets["summary_grade"]

            # Define Styles
            s_header_fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
            s_header_font = Font(color="FFFFFF", bold=True)
            center_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
            left_align = Alignment(horizontal="left", vertical="center", wrap_text=True)
            thin_border = Border(
                left=Side(style="thin", color="BFBFBF"),
                right=Side(style="thin", color="BFBFBF"),
                top=Side(style="thin", color="BFBFBF"),
                bottom=Side(style="thin", color="BFBFBF"),
            )

            # Grade Colors Map
            grade_colors = {
                "A": PatternFill("solid", fgColor="C6EFCE"),  # Green
                "B": PatternFill("solid", fgColor="B4C6E7"),  # Light Blue
                "C": PatternFill("solid", fgColor="FFEB9C"),  # Yellow
                "D": PatternFill("solid", fgColor="FFC7CE"),  # Red
            }

            # Apply Styles to Headers
            for col_num, cell in enumerate(worksheet[1], 1):
                cell.fill = s_header_fill
                cell.font = s_header_font
                cell.alignment = center_align
                cell.border = thin_border

            # Find the index of the "Avg_eva_grade" column to apply color
            grade_col_idx = list(summary_df.columns).index("Avg_eva_grade")

            # Apply Styles to Data Rows and Adjust Row Heights
            for row in worksheet.iter_rows(
                min_row=2,
                max_row=worksheet.max_row,
                min_col=1,
                max_col=worksheet.max_column,
            ):
                worksheet.row_dimensions[row[0].row].height = 20  # Set standard row height
                
                for idx, cell in enumerate(row):
                    cell.border = thin_border
                    
                    if idx in [0, 1, 2]:  # Email, Name, Team Name
                        cell.alignment = left_align
                    else:
                        cell.alignment = center_align

                    # Apply Color to the Grade Column
                    if idx == grade_col_idx:
                        grade_val = str(cell.value)
                        if grade_val in grade_colors:
                            cell.fill = grade_colors[grade_val]

            # Better Column Auto-fit (based on max string length in each column)
            for col_idx in range(1, worksheet.max_column + 1):
                col_letter = get_column_letter(col_idx)
                max_len = 0
                for cell in worksheet[col_letter]:
                    if cell.value:
                        max_len = max(max_len, len(str(cell.value)))
                worksheet.column_dimensions[col_letter].width = max(max_len + 2, 12)

        log.info("summary_grade_sheet_added_with_formatting", path=output_path)

    except Exception as exc:
        log.error("summary_grade_sheet_failed", error=str(exc))

    return str(output_path)