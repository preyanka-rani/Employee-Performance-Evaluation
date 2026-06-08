"""
app/teams/cirt_infra/report.py
──────────────────────────────
Generates the formatted Excel report for CIRT & Infra Team evaluation runs.

Sheet 1 – "Detailed Report"
    Identity: Employee ID | Name | Email | <team col> | year | month
    Segment A: Total Log Hours | Log Hours Score | Sentiment Score |
               Monthly Functional Score | Segment A Marks (0-30)
    Segment B: Attendance Score | Attendance Marks (0-10) |
               <TL readiness col> | <TL kpi col> | <TL general col> |
               TL Total (0-40) | Segment B Marks (0-50)
    Final:     Base Total (0-80) | Final Score (0-100)

Sheet 2 – "Final Summary" (EBS of Business Automation Ltd. format)
    Team | Employee | Email | avg_functional_score |
    <TL readiness col - original name> |
    avg_office_discipline_score |
    <TL kpi col - original name> |
    <TL general col - original name> |
    Avg_eval_scores | Percentage | Avg_eva_grade

Grade (applied to Percentage × 100):
    88–100 → A  |  84–87 → B  |  75–83 → C  |  0–74 → D

Saved to: outputs/cirt_infra/CIRT_Infra_Final_Report_{team_key}_{year}_{month:02d}.xlsx
"""

from __future__ import annotations

import pathlib
from typing import Any

import openpyxl
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from sqlalchemy import select
from sqlalchemy.ext.asyncio import AsyncSession

from app.core.logging_config import get_logger
from app.models.scores import (
    AttendanceScore,
    FinalScore,
    SentimentScore,
    WorkLogScore,
)
from app.repositories.employee_repository import EmployeeRepository

logger = get_logger(__name__)

# Output directory – CIRT & Infra reports live here
_REPORTS_DIR = pathlib.Path("outputs") / "cirt_infra"

# ── Fallback column names (used when original headers are not available) ──────
_DEFAULT_COL_NAMES: dict[str, str] = {
    "team_name": "team_name",
    "support_readiness": "Support Readiness & Issue Handling (0-10)",
    "kpi": "KPI Agreement (0-15)",
    "general": "Leadership General Assessment (0-15)",
}


# ── Grade helper ──────────────────────────────────────────────────────────────


def _get_grade_from_pct(pct: float) -> str:
    """
    Return letter grade from a percentage value (0-100 scale).
    Thresholds match the EBS of Business Automation Ltd. format.
    """
    if pct >= 88:
        return "A"
    if pct >= 84:
        return "B"
    if pct >= 75:
        return "C"
    return "D"


# ── Data fetch helper ─────────────────────────────────────────────────────────


async def _load_score_rows(
    db: AsyncSession,
    run_id: int,
    emails: list[str],
) -> dict[str, dict[str, Any]]:
    """
    Fetch all score rows for the run, indexed by employee_email.

    Returns a dict keyed by email; each value is a dict with the CIRT
    fields needed by the report (so we only touch the DB once).
    """
    if not emails:
        return {}

    # WorkLogScore
    wl_rows = (
        await db.execute(
            select(WorkLogScore).where(
                WorkLogScore.evaluation_run_id == run_id,
                WorkLogScore.employee_email.in_(emails),
            )
        )
    ).scalars().all()

    # SentimentScore
    sent_rows = (
        await db.execute(
            select(SentimentScore).where(
                SentimentScore.evaluation_run_id == run_id,
                SentimentScore.employee_email.in_(emails),
            )
        )
    ).scalars().all()

    # AttendanceScore
    att_rows = (
        await db.execute(
            select(AttendanceScore).where(
                AttendanceScore.evaluation_run_id == run_id,
                AttendanceScore.employee_email.in_(emails),
            )
        )
    ).scalars().all()

    # FinalScore
    fs_rows = (
        await db.execute(
            select(FinalScore).where(
                FinalScore.evaluation_run_id == run_id,
                FinalScore.employee_email.in_(emails),
            )
        )
    ).scalars().all()

    by_email: dict[str, dict[str, Any]] = {}
    for fs in fs_rows:
        by_email[fs.employee_email] = {
            "total_log_hours": 0.0,
            "log_hours_score": 0.0,
            "sentiment_score": 0.0,
            "avg_polarity": 0.0,
            "total_logs_analyzed": 0,
            "attendance_score": float(fs.attendance_score or 0.0),
            "attendance_marks": float(fs.attendance_marks or 0.0),
            "support_readiness": float(fs.problem_solving or 0.0),
            "kpi": float(fs.kpi or 0.0),
            "general": float(fs.general_assessment or 0.0),
            "tl_total": float(fs.tl_total or 0.0),
            "segment_a_marks": float(fs.segment_a_marks or 0.0),
            "segment_b_marks": float(fs.segment_b_marks or 0.0),
            "base_total": float(fs.base_total or 0.0),
            "final_score": float(fs.final_score or 0.0),
        }

    for wl in wl_rows:
        rec = by_email.get(wl.employee_email)
        if rec is None:
            continue
        rec["total_log_hours"] = float(wl.total_log_hours or 0.0)
        rec["log_hours_score"] = float(wl.normalized_score or 0.0)

    for s in sent_rows:
        rec = by_email.get(s.employee_email)
        if rec is None:
            continue
        rec["sentiment_score"] = float(s.score or 0.0)
        rec["avg_polarity"] = float(s.average_polarity or 0.0)
        rec["total_logs_analyzed"] = int(s.total_logs_analyzed or 0)

    for a in att_rows:
        rec = by_email.get(a.employee_email)
        if rec is None:
            continue
        # The FinalScore row already carries attendance_score; only update
        # the present_days/late counters here if needed.
        rec.setdefault("present_days", int(a.present_days or 0))
        rec.setdefault("late_attendance", int(a.late_attendance or 0))

    return by_email


# ── Public entry point ────────────────────────────────────────────────────────


async def generate_cirt_infra_excel_report(
    run_id: int,
    emails: list[str],
    team: str,
    year: int,
    month: int,
    db: AsyncSession,
    col_names: dict[str, str] | None = None,
    team_display_name: str = "",
) -> str:
    """
    Build a formatted two-sheet Excel report for a completed CIRT & Infra
    Team bulk evaluation run.

    Args:
        col_names: canonical→original_header mapping from the uploaded Excel.
                   Used to preserve original column names in both sheets.
        team_display_name: raw team name from the uploaded Excel.  Falls
                           back to ``team`` key.

    Returns:
        Absolute file path to the saved .xlsx report.
    """
    # ── Resolve dynamic column names ──────────────────────────────────────────
    _cn = col_names or {}
    team_col_hdr = _cn.get("team_name", _DEFAULT_COL_NAMES["team_name"])
    tl_readiness_hdr = _cn.get(
        "support_readiness", _DEFAULT_COL_NAMES["support_readiness"]
    )
    tl_kpi_hdr = _cn.get("kpi", _DEFAULT_COL_NAMES["kpi"])
    tl_general_hdr = _cn.get("general", _DEFAULT_COL_NAMES["general"])
    team_display = team_display_name or team

    # ── Fetch scores ──────────────────────────────────────────────────────────
    by_email = await _load_score_rows(db, run_id=run_id, emails=emails)

    # ── Employee lookup ───────────────────────────────────────────────────────
    emp_repo = EmployeeRepository(db)
    email_to_name: dict[str, str] = {}
    email_to_id: dict[str, str] = {}
    for email in emails:
        emp = await emp_repo.get_by_email(email)
        if emp is not None:
            email_to_name[email] = emp.name or email
            email_to_id[email] = emp.employee_id or ""

    # Sort descending by final_score for consistent ordering
    scores_sorted = sorted(
        by_email.values(),
        key=lambda r: r["final_score"],
        reverse=True,
    )

    # ── Common fill styles ────────────────────────────────────────────────────
    hdr_font = Font(bold=True, color="FFFFFF")
    hdr_fill = PatternFill("solid", fgColor="2E4057")
    hdr_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    center_align = Alignment(horizontal="center", vertical="center")

    green_fill = PatternFill("solid", fgColor="C6EFCE")  # ≥ 80
    yellow_fill = PatternFill("solid", fgColor="FFEB9C")  # 60-79
    red_fill = PatternFill("solid", fgColor="FFC7CE")  # < 60
    grade_a_fill = PatternFill("solid", fgColor="00B050")  # A – dark green
    grade_b_fill = PatternFill("solid", fgColor="92D050")  # B – light green
    grade_c_fill = PatternFill("solid", fgColor="FFEB9C")  # C – yellow
    grade_d_fill = PatternFill("solid", fgColor="FFC7CE")  # D – red
    section_fill = PatternFill("solid", fgColor="D6E4F0")  # section header rows
    total_fill = PatternFill("solid", fgColor="1F4E79")

    thin_side = Side(style="thin", color="AAAAAA")
    thin_border = Border(
        left=thin_side, right=thin_side, top=thin_side, bottom=thin_side
    )

    # ═════════════════════════════════════════════════════════════════════════
    # SHEET 1 – Detailed Report
    # ═════════════════════════════════════════════════════════════════════════
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Detailed Report"

    detail_headers = [
        "Employee ID",
        "Name",
        "Email",
        team_col_hdr,  # original header from input Excel
        "year",
        "month",
        # ── Segment A: CRM Log + Sentiment ──
        "Total Log Hours",
        "Log Hours Score (0-100)",
        "Sentiment Score (0-100)",
        "Monthly Functional Score (0-100)",
        "Segment A Marks (0-30)",
        # ── Segment B ──
        "Attendance Score (0-100)",
        "Attendance Marks (0-10)",
        tl_readiness_hdr,
        tl_kpi_hdr,
        tl_general_hdr,
        "TL Total (0-40)",
        "Segment B Marks (0-50)",
        # ── Final ──
        "Base Total (0-80)",
        "Final Score (0-100)",
    ]

    detail_widths = [
        14,
        22,
        30,
        max(18, len(team_col_hdr) + 2),
        8,
        8,
        18,  # Total Log Hours
        22,  # Log Hours Score
        22,  # Sentiment Score
        28,  # Monthly Functional Score
        22,  # Segment A Marks
        22,  # Attendance Score
        20,  # Attendance Marks
        max(24, len(tl_readiness_hdr) + 2),
        max(20, len(tl_kpi_hdr) + 2),
        max(24, len(tl_general_hdr) + 2),
        16,  # TL Total
        22,  # Segment B Marks
        18,  # Base Total
        18,  # Final Score
    ]

    ws.append(detail_headers)
    ws.row_dimensions[1].height = 50
    for col_idx in range(1, len(detail_headers) + 1):
        cell = ws.cell(row=1, column=col_idx)
        cell.font = hdr_font
        cell.fill = hdr_fill
        cell.alignment = hdr_align
        cell.border = thin_border

    ws.freeze_panes = "A2"

    for col_idx, width in enumerate(detail_widths, start=1):
        ws.column_dimensions[get_column_letter(col_idx)].width = width

    for rec in scores_sorted:
        emp_email = next(
            (e for e, r in by_email.items() if r is rec), ""
        )
        if not emp_email:
            # Fallback – match by tuple identity won't work post-sort; rebuild
            for e, r in by_email.items():
                if r is rec:
                    emp_email = e
                    break

        row_data = [
            email_to_id.get(emp_email, ""),
            email_to_name.get(emp_email, emp_email),
            emp_email,
            team_display,
            year,
            month,
            # Segment A
            round(rec["total_log_hours"], 2),
            round(rec["log_hours_score"], 2),
            round(rec["sentiment_score"], 2),
            round(rec["segment_a_marks"] / 0.30 if rec["segment_a_marks"] else 0.0, 2),
            round(rec["segment_a_marks"], 2),
            # Segment B
            round(rec["attendance_score"], 2),
            round(rec["attendance_marks"], 2),
            round(rec["support_readiness"], 2),
            round(rec["kpi"], 2),
            round(rec["general"], 2),
            round(rec["tl_total"], 2),
            round(rec["segment_b_marks"], 2),
            # Final
            round(rec["base_total"], 2),
            round(rec["final_score"], 2),
        ]
        ws.append(row_data)

        row_num = ws.max_row
        for col_idx in range(1, len(detail_headers) + 1):
            c = ws.cell(row=row_num, column=col_idx)
            c.alignment = center_align
            c.border = thin_border

        # Color-code Final Score (last column)
        final_cell = ws.cell(row=row_num, column=len(detail_headers))
        final_score = rec["final_score"]
        if final_score >= 80.0:
            final_cell.fill = green_fill
        elif final_score >= 60.0:
            final_cell.fill = yellow_fill
        else:
            final_cell.fill = red_fill

    # Team-average footer row
    if scores_sorted:
        ws.append([])
        avg_final = round(
            sum(r["final_score"] for r in scores_sorted) / len(scores_sorted), 2
        )
        summary = [""] * len(detail_headers)
        summary[0] = "TEAM AVERAGE"
        # Reconstruct the monthly_functional_score column from segment_a_marks
        avg_monthly_func = round(
            sum(
                (r["segment_a_marks"] / 0.30) if r["segment_a_marks"] else 0.0
                for r in scores_sorted
            )
            / len(scores_sorted),
            2,
        )
        avg_seg_a = round(
            sum(r["segment_a_marks"] for r in scores_sorted) / len(scores_sorted), 2
        )
        avg_seg_b = round(
            sum(r["segment_b_marks"] for r in scores_sorted) / len(scores_sorted), 2
        )
        avg_base = round(
            sum(r["base_total"] for r in scores_sorted) / len(scores_sorted), 2
        )
        # summary indices:  0  1  2  3  4  5 | 6 7 8 9 10 | 11 12 13 14 15 16 17 | 18 19
        summary[6] = round(
            sum(r["total_log_hours"] for r in scores_sorted) / len(scores_sorted), 2
        )
        summary[7] = round(
            sum(r["log_hours_score"] for r in scores_sorted) / len(scores_sorted), 2
        )
        summary[8] = round(
            sum(r["sentiment_score"] for r in scores_sorted) / len(scores_sorted), 2
        )
        summary[9] = avg_monthly_func
        summary[10] = avg_seg_a
        summary[11] = round(
            sum(r["attendance_score"] for r in scores_sorted) / len(scores_sorted), 2
        )
        summary[12] = round(
            sum(r["attendance_marks"] for r in scores_sorted) / len(scores_sorted), 2
        )
        summary[13] = round(
            sum(r["support_readiness"] for r in scores_sorted) / len(scores_sorted), 2
        )
        summary[14] = round(
            sum(r["kpi"] for r in scores_sorted) / len(scores_sorted), 2
        )
        summary[15] = round(
            sum(r["general"] for r in scores_sorted) / len(scores_sorted), 2
        )
        summary[16] = round(
            sum(r["tl_total"] for r in scores_sorted) / len(scores_sorted), 2
        )
        summary[17] = avg_seg_b
        summary[18] = avg_base
        summary[19] = avg_final
        ws.append(summary)
        last_row = ws.max_row
        for col_idx in range(1, len(detail_headers) + 1):
            c = ws.cell(row=last_row, column=col_idx)
            c.font = Font(bold=True)
            c.border = thin_border
        avg_cell = ws.cell(row=last_row, column=len(detail_headers))
        avg_cell.fill = (
            green_fill
            if avg_final >= 80
            else yellow_fill if avg_final >= 60 else red_fill
        )

    # ═════════════════════════════════════════════════════════════════════════
    # SHEET 2 – Final Summary (EBS of Business Automation Ltd. format)
    # ═════════════════════════════════════════════════════════════════════════
    ws2 = wb.create_sheet("Final Summary")

    summary_headers = [
        "Team",
        "Employee",
        "Email",
        "avg_functional_score",
        tl_readiness_hdr,
        "avg_office_discipline_score",
        tl_kpi_hdr,
        tl_general_hdr,
        "Avg_eval_scores",
        "Percentage",
        "Avg_eva_grade",
    ]

    ws2.append(summary_headers)
    ws2.row_dimensions[1].height = 50
    for col_idx in range(1, len(summary_headers) + 1):
        cell = ws2.cell(row=1, column=col_idx)
        cell.font = hdr_font
        cell.fill = hdr_fill
        cell.alignment = hdr_align
        cell.border = thin_border

    ws2.freeze_panes = "A2"

    summary_col_widths = [
        max(18, len(team_display) + 2),
        24,
        32,
        20,
        max(24, len(tl_readiness_hdr) + 2),
        26,
        max(20, len(tl_kpi_hdr) + 2),
        max(24, len(tl_general_hdr) + 2),
        18,
        12,
        14,
    ]
    for col_idx, width in enumerate(summary_col_widths, start=1):
        ws2.column_dimensions[get_column_letter(col_idx)].width = width

    grade_fills = {
        "A": grade_a_fill,
        "B": grade_b_fill,
        "C": grade_c_fill,
        "D": grade_d_fill,
    }

    _MAX_SCORE = 80.0  # CIRT & Infra base maximum

    for rec in scores_sorted:
        emp_email = next((e for e, r in by_email.items() if r is rec), "")
        if not emp_email:
            for e, r in by_email.items():
                if r is rec:
                    emp_email = e
                    break

        # avg_eval_scores = segment_a_marks + support_readiness + attendance_marks + kpi + general
        avg_eval = round(
            rec["segment_a_marks"]
            + rec["support_readiness"]
            + rec["attendance_marks"]
            + rec["kpi"]
            + rec["general"],
            2,
        )
        percentage = round(avg_eval / _MAX_SCORE, 2)
        grade = _get_grade_from_pct(round(percentage * 100, 2))

        ws2.append(
            [
                team_display,
                email_to_name.get(emp_email, emp_email),
                emp_email,
                round(rec["segment_a_marks"], 2),
                round(rec["support_readiness"], 2),
                round(rec["attendance_marks"], 2),
                round(rec["kpi"], 2),
                round(rec["general"], 2),
                avg_eval,
                percentage,
                grade,
            ]
        )

        row_num = ws2.max_row
        for col_idx in range(1, len(summary_headers) + 1):
            c = ws2.cell(row=row_num, column=col_idx)
            c.alignment = center_align
            c.border = thin_border

        # Colour-code grade cell (column 11)
        grade_cell = ws2.cell(row=row_num, column=11)
        grade_cell.font = Font(
            bold=True, color="FFFFFF" if grade in ("A", "B") else "000000"
        )
        grade_cell.fill = grade_fills[grade]

        # Colour Percentage cell (column 10)
        pct_cell = ws2.cell(row=row_num, column=10)
        pct_val = percentage * 100
        pct_cell.fill = (
            grade_a_fill
            if pct_val >= 88
            else (
                grade_b_fill
                if pct_val >= 84
                else grade_c_fill if pct_val >= 75 else grade_d_fill
            )
        )

    # Team-average footer row
    if scores_sorted:
        ws2.append([])
        avg_seg_a = round(
            sum(r["segment_a_marks"] for r in scores_sorted) / len(scores_sorted), 2
        )
        avg_readiness = round(
            sum(r["support_readiness"] for r in scores_sorted) / len(scores_sorted), 2
        )
        avg_attendance = round(
            sum(r["attendance_marks"] for r in scores_sorted) / len(scores_sorted), 2
        )
        avg_kpi = round(sum(r["kpi"] for r in scores_sorted) / len(scores_sorted), 2)
        avg_general = round(
            sum(r["general"] for r in scores_sorted) / len(scores_sorted), 2
        )
        avg_eval_all = round(
            avg_seg_a + avg_readiness + avg_attendance + avg_kpi + avg_general, 2
        )
        avg_pct = round(avg_eval_all / _MAX_SCORE, 2)
        avg_grade = _get_grade_from_pct(round(avg_pct * 100, 2))

        ws2.append(
            [
                "TEAM AVERAGE",
                "",
                "",
                avg_seg_a,
                avg_readiness,
                avg_attendance,
                avg_kpi,
                avg_general,
                avg_eval_all,
                avg_pct,
                avg_grade,
            ]
        )
        last_row = ws2.max_row
        for col_idx in range(1, len(summary_headers) + 1):
            c = ws2.cell(row=last_row, column=col_idx)
            c.font = Font(bold=True, color="FFFFFF" if col_idx == 1 else "000000")
            c.fill = (
                PatternFill("solid", fgColor="1F4E79") if col_idx == 1 else section_fill
            )
            c.border = thin_border
            c.alignment = center_align

    # ── Save ──────────────────────────────────────────────────────────────────
    _REPORTS_DIR.mkdir(parents=True, exist_ok=True)
    filename = f"CIRT_Infra_Final_Report_{team}_{year}_{month:02d}.xlsx"
    output_path = _REPORTS_DIR / filename
    wb.save(str(output_path))

    logger.info(
        "cirt_infra_report_saved",
        path=str(output_path),
        employee_count=len(scores_sorted),
        team=team,
        year=year,
        month=month,
    )

    return str(output_path.resolve())
