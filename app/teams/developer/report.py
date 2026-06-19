"""
app/teams/developer/report.py
─────────────────────────────
Developer worker report generator with exact summary_grade formatting.

Thin wrapper over ``app.services.reporting.report_generator`` — the
legacy functions already save their outputs to the canonical
``outputs/developer/`` directory, so this module simply resolves the
absolute paths and surfaces them in the contract's expected shape.

It also automatically appends a beautifully formatted `summary_grade` 
sheet perfectly matching the structure of `developer_04_2026.xlsx`.
"""

from __future__ import annotations

import os
import pathlib
import pandas as pd

from sqlalchemy.ext.asyncio import AsyncSession
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

from app.core.logging_config import get_logger
from app.services.reporting import report_generator
from app.teams.developer.formulas import calculate_evaluation_grade

logger = get_logger(__name__)


async def generate_developer_reports(
    *,
    run_id: int,
    emails: list[str],
    year: int,
    month: int,
    db: AsyncSession,
) -> dict[str, str]:
    """
    Generate both developer Excel reports and append the formatted summary_grade sheet.
    """
    log = logger.bind(run_id=run_id, year=year, month=month, count=len(emails))
    log.info("developer_reports_start")

    # 1. Code-quality report (per-project analysis)
    cq_path = await report_generator.generate_code_quality_report(
        run_id=run_id,
        emails=emails,
        team="developer",
        year=year,
        month=month,
        db=db,
    )

    # 2. Final score report (24-column component breakdown)
    fr_path = await report_generator.generate_excel_report(
        run_id=run_id,
        emails=emails,
        team="developer",
        year=year,
        month=month,
        db=db,
    )

    # 3. Post-process: Automatically append the 'summary_grade' sheet with nice formatting
    if fr_path and os.path.exists(fr_path):
        try:
            # Read the generated final report (main sheet)
            df = pd.read_excel(fr_path, sheet_name=0)
            
            # Initialize exact matching dataframe structure
            summary_df = pd.DataFrame()
            
            # Map Base Info
            summary_df["emp_email"] = df.get("Email", "")
            summary_df["emp_name"] = df.get("Name", "")
            summary_df["team_name"] = df.get("Team", "")
            
            # Map Segment A and Segment B Data
            summary_df["avg_functional_job_performance_50"] = df.get("Segment A Marks (0-50)", 0.0)
            summary_df["avg_office_discipline_10"] = df.get("Attendance Score (0-100)", 0.0) / 10.0
            summary_df["avg_critical_thinking_and_problem_solving_10"] = df.get("TL Problem Solving (0-10)", 0.0)
            summary_df["avg_monthly_performance_agreement_15"] = df.get("TL KPI (0-15)", 0.0)
            summary_df["avg_team_leader_assessment_15"] = df.get("TL General (0-15)", 0.0)
            
            # Calculate Base Total (Out of 100 for Developers)
            summary_df["avg_total_scores"] = (
                summary_df["avg_functional_job_performance_50"] +
                summary_df["avg_office_discipline_10"] +
                summary_df["avg_critical_thinking_and_problem_solving_10"] +
                summary_df["avg_monthly_performance_agreement_15"] +
                summary_df["avg_team_leader_assessment_15"]
            )
            
            # Map Reward Score
            summary_df["reword_score_5"] = df.get("Reward Score (0-5)", 0.0)
            
            # Finalize Score and Percentage calculations
            summary_df["finalize_score"] = df.get("Final Score", 0.0)
            summary_df["score_percentage"] = (summary_df["finalize_score"] / 100.0).round(4)
            
            # Apply Grading Rules using the imported function
            summary_df["Avg_eva_grade"] = summary_df["finalize_score"].apply(calculate_evaluation_grade)
            
            # Append the summary_grade sheet to the existing Excel workbook
            with pd.ExcelWriter(fr_path, engine='openpyxl', mode='a', if_sheet_exists='replace') as writer:
                summary_df.to_excel(writer, sheet_name='summary_grade', index=False)
                
                # --- START OF BEAUTIFUL FORMATTING ---
                workbook = writer.book
                worksheet = writer.sheets['summary_grade']
                
                # Define Styles matching your image
                header_fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid") # Dark Blue
                header_font = Font(color="FFFFFF", bold=True) # White & Bold
                center_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
                left_align = Alignment(horizontal="left", vertical="center", wrap_text=True)
                thin_border = Border(
                    left=Side(style='thin', color="BFBFBF"), 
                    right=Side(style='thin', color="BFBFBF"),
                    top=Side(style='thin', color="BFBFBF"), 
                    bottom=Side(style='thin', color="BFBFBF")
                )

                # Format Headers (Row 1)
                for col_num, cell in enumerate(worksheet[1], 1):
                    cell.fill = header_fill
                    cell.font = header_font
                    cell.alignment = center_align
                    cell.border = thin_border
                    
                    # Adjust column widths dynamically based on header text
                    col_letter = get_column_letter(col_num)
                    header_length = len(str(cell.value))
                    worksheet.column_dimensions[col_letter].width = max(header_length + 2, 12)

                # Format Data Rows
                for row in worksheet.iter_rows(min_row=2, max_row=worksheet.max_row, min_col=1, max_col=worksheet.max_column):
                    for idx, cell in enumerate(row):
                        cell.border = thin_border
                        # Make Emails and Names left-aligned for better readability, everything else centered
                        if idx in [0, 1, 2]: # emp_email, emp_name, team_name
                            cell.alignment = left_align
                        else:
                            cell.alignment = center_align
                # --- END OF BEAUTIFUL FORMATTING ---

            log.info("summary_grade_sheet_added_with_formatting", path=fr_path)
            
        except Exception as exc:
            log.error("summary_grade_sheet_failed", error=str(exc))

    log.info(
        "developer_reports_done",
        code_quality=cq_path,
        final=fr_path,
    )
    return {
        "code_quality_report": str(pathlib.Path(cq_path).resolve()) if cq_path else "",
        "final_report": str(pathlib.Path(fr_path).resolve()) if fr_path else "",
    }